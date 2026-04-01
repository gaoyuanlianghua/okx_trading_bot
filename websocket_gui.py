"""
WebSocket 客户端界面

使用 PyQt5 创建一个基于 WebSocket 客户端的实时交易界面
"""

import sys
import asyncio
import json
import logging
import os
from datetime import datetime
from cryptography.fernet import Fernet
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QGridLayout,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QLabel,
    QLineEdit,
    QPushButton,
    QComboBox,
    QStatusBar,
    QSplitter,
    QGroupBox,
    QFormLayout,
    QFrame,
    QMessageBox,
    QSizePolicy,
    QTextEdit,
    QDialog,
    QCheckBox,
)
from PyQt5.QtCore import QPropertyAnimation, QParallelAnimationGroup, QEasingCurve, QTimer, pyqtSignal, Qt, QSize, QMimeData
from PyQt5.QtGui import QColor, QFont, QIcon, QDrag

# 导入图表库
import matplotlib

matplotlib.use("Qt5Agg")
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import mplfinance as mpf
import numpy as np
import pandas as pd

from core import OKXWebSocketClient, EventBus, EventType
from core.monitoring import strategy_monitor
from core.storage.data_persistence import data_persistence
from core.agents.market_data_agent import MarketDataAgent
from core.agents.base_agent import AgentConfig

# 设置日志
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


class DraggableDashboardCard(QGroupBox):
    """
    可拖拽的仪表盘卡片
    """
    def __init__(self, title, parent=None):
        super().__init__(title, parent)
        self.setAcceptDrops(True)
        self.setMinimumHeight(160)
        self.setStyleSheet("QGroupBox { border-radius: 10px; }")
    
    def mousePressEvent(self, event):
        """
        鼠标按下事件，开始拖拽
        """
        if event.button() == Qt.LeftButton:
            # 创建拖拽对象
            drag = QDrag(self)
            mime_data = QMimeData()
            mime_data.setText(self.title())
            drag.setMimeData(mime_data)
            
            # 启动拖拽
            drag.exec_(Qt.MoveAction)
    
    def dragEnterEvent(self, event):
        """
        拖拽进入事件
        """
        if event.mimeData().hasText():
            event.acceptProposedAction()
    
    def dropEvent(self, event):
        """
        拖拽释放事件
        """
        if event.mimeData().hasText():
            # 获取拖拽的卡片标题
            dragged_card_title = event.mimeData().text()
            
            # 交换位置逻辑
            # 这里需要获取父布局，然后交换两个卡片的位置
            parent_layout = self.parent().layout()
            if parent_layout:
                # 找到被拖拽的卡片和目标卡片
                dragged_card = None
                target_card = self
                
                # 遍历布局中的所有widget
                for i in range(parent_layout.count()):
                    widget = parent_layout.itemAt(i).widget()
                    if isinstance(widget, DraggableDashboardCard) and widget.title() == dragged_card_title:
                        dragged_card = widget
                        break
                
                if dragged_card and dragged_card != target_card:
                    # 保存两个卡片的位置
                    dragged_index = parent_layout.indexOf(dragged_card)
                    target_index = parent_layout.indexOf(target_card)
                    
                    # 交换位置
                    if dragged_index < target_index:
                        parent_layout.insertWidget(target_index, dragged_card)
                        parent_layout.insertWidget(dragged_index, target_card)
                    else:
                        parent_layout.insertWidget(dragged_index, target_card)
                        parent_layout.insertWidget(target_index, dragged_card)
                    
                event.acceptProposedAction()

class AnimatedTabWidget(QTabWidget):
    """
    带动画效果的标签页组件
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.currentChanged.connect(self._on_tab_changed)
        
    def _on_tab_changed(self, index):
        """
        标签页切换时的动画效果
        """
        try:
            # 获取当前标签页
            current_widget = self.widget(index)
            if not current_widget:
                return
            
            # 创建淡入动画
            fade_in = QPropertyAnimation(current_widget, b"windowOpacity")
            fade_in.setDuration(300)
            fade_in.setStartValue(0.0)
            fade_in.setEndValue(1.0)
            fade_in.setEasingCurve(QEasingCurve.InOutQuad)
            
            # 启动动画
            fade_in.start()
        except Exception as e:
            logger.error(f"标签页动画错误: {e}")

class WebSocketGUI(QMainWindow):
    """
    WebSocket 客户端界面
    """

    # 信号定义
    update_market_data = pyqtSignal(dict)
    update_order_data = pyqtSignal(dict)
    update_account_data = pyqtSignal(dict)
    update_connection_status = pyqtSignal(bool, str)

    def __init__(self):
        """
        初始化 GUI
        """
        super().__init__()

        # 初始化权限控制
        self._init_permission_system()
        
        # 初始化事件总线
        self.event_bus = EventBus()

        # 数据存储
        self.market_data = {}
        # 价格历史数据，用于图表显示
        self.price_history = {}  # 格式: {inst_id: [(timestamp, price), ...]}
        # 加载本地缓存的订单历史
        self.order_data = data_persistence.load_order_history()
        # 加载本地缓存的市场数据
        self._load_market_data_cache()
        self.account_data = {}
        self.strategies = []
        self.strategy_instances = {}  # 存储策略实例
        self.strategy_threads = {}  # 存储策略线程
        
        # 初始化市场数据智能体
        self.market_data_agent = None
        # 初始化WebSocket客户端
        self.ws_client = None

        # 线程池管理
        import concurrent.futures

        self._thread_pool = concurrent.futures.ThreadPoolExecutor(
            max_workers=5
        )  # 限制线程池大小

        # 初始化加密密钥
        self._init_encryption()
        
        # 加载加密的API密钥
        self._load_encrypted_api_keys()

        # 加载现有策略
        self._load_strategies()

        # 验证用户登录
        if not self._show_login_dialog():
            sys.exit()

        # 初始化界面
        self.init_ui()

        # 连接信号
        self.update_market_data.connect(self.on_market_data_update)
        self.update_order_data.connect(self.on_order_data_update)
        self.update_account_data.connect(self.on_account_data_update)
        self.update_connection_status.connect(self.on_connection_status_update)

        # 启动事件总线
        self.event_bus.start()

        # 注册事件监听器
        self.event_bus.subscribe(EventType.MARKET_DATA_TICKER, self.on_market_event)
        self.event_bus.subscribe(EventType.ORDER_UPDATED, self.on_order_event)
        self.event_bus.subscribe(EventType.CUSTOM, self.on_account_event)
        self.event_bus.subscribe(EventType.WS_CONNECTED, self.on_ws_connected)
        self.event_bus.subscribe(EventType.WS_DISCONNECTED, self.on_ws_disconnected)

    def _init_permission_system(self):
        """
        初始化权限系统
        """
        # 定义用户角色和权限
        self.roles = {
            "admin": {
                "name": "管理员",
                "permissions": [
                    "view_market_data",
                    "view_account",
                    "place_order",
                    "manage_strategies",
                    "manage_users",
                    "modify_settings"
                ]
            },
            "trader": {
                "name": "交易员",
                "permissions": [
                    "view_market_data",
                    "view_account",
                    "place_order",
                    "manage_strategies"
                ]
            },
            "viewer": {
                "name": "查看者",
                "permissions": [
                    "view_market_data",
                    "view_account"
                ]
            }
        }
        
        # 默认用户（实际应用中应该从数据库或配置文件加载）
        self.users = {
            "admin": {
                "password": "admin123",  # 实际应用中应该使用哈希密码
                "role": "admin"
            },
            "trader": {
                "password": "trader123",
                "role": "trader"
            },
            "viewer": {
                "password": "viewer123",
                "role": "viewer"
            }
        }
        
        # 当前登录用户
        self.current_user = None
        self.current_role = None

    def _show_login_dialog(self):
        """
        显示登录对话框
        """
        dialog = QDialog(self)
        dialog.setWindowTitle("登录")
        dialog.setGeometry(300, 300, 400, 200)
        
        layout = QVBoxLayout(dialog)
        
        # 用户名输入
        user_layout = QHBoxLayout()
        user_layout.addWidget(QLabel("用户名:"))
        user_input = QLineEdit()
        user_input.setPlaceholderText("请输入用户名")
        user_layout.addWidget(user_input)
        layout.addLayout(user_layout)
        
        # 密码输入
        pass_layout = QHBoxLayout()
        pass_layout.addWidget(QLabel("密码:"))
        pass_input = QLineEdit()
        pass_input.setPlaceholderText("请输入密码")
        pass_input.setEchoMode(QLineEdit.Password)
        pass_layout.addWidget(pass_input)
        layout.addLayout(pass_layout)
        
        # 错误信息
        error_label = QLabel("")
        error_label.setStyleSheet("color: red;")
        layout.addWidget(error_label)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        cancel_button = QPushButton("取消")
        login_button = QPushButton("登录")
        
        button_layout.addStretch(1)
        button_layout.addWidget(cancel_button)
        button_layout.addWidget(login_button)
        layout.addLayout(button_layout)
        
        # 登录结果
        login_success = False
        
        def on_login():
            nonlocal login_success
            username = user_input.text().strip()
            password = pass_input.text().strip()
            
            if username in self.users and self.users[username]["password"] == password:
                self.current_user = username
                self.current_role = self.users[username]["role"]
                login_success = True
                dialog.accept()
            else:
                error_label.setText("用户名或密码错误")
        
        def on_cancel():
            dialog.reject()
        
        login_button.clicked.connect(on_login)
        cancel_button.clicked.connect(on_cancel)
        
        # 显示对话框
        result = dialog.exec_()
        
        if login_success:
            logger.info(f"用户 {self.current_user} (角色: {self.roles[self.current_role]['name']}) 登录成功")
            return True
        else:
            logger.warning("用户登录失败")
            return False

    def has_permission(self, permission):
        """
        检查用户是否有指定权限
        """
        if not self.current_role:
            return False
        
        role_permissions = self.roles.get(self.current_role, {}).get("permissions", [])
        return permission in role_permissions

    def update_ui_permissions(self):
        """
        根据用户权限更新UI元素的可用性
        """
        # 交易操作权限
        can_trade = self.has_permission("place_order")
        self.execute_trade_button.setEnabled(can_trade)
        self.execute_batch_trade_button.setEnabled(can_trade)
        self.execute_conditional_trade_button.setEnabled(can_trade)
        
        # 策略管理权限
        can_manage_strategies = self.has_permission("manage_strategies")
        self.create_strategy_button.setEnabled(can_manage_strategies)
        self.start_strategy_button.setEnabled(can_manage_strategies)
        self.stop_strategy_button.setEnabled(can_manage_strategies)
        self.edit_strategy_button.setEnabled(can_manage_strategies)
        self.delete_strategy_button.setEnabled(can_manage_strategies)
        
        # API配置权限
        can_modify_settings = self.has_permission("modify_settings")
        self.api_key_input.setEnabled(can_modify_settings)
        self.secret_input.setEnabled(can_modify_settings)
        self.passphrase_input.setEnabled(can_modify_settings)
        self.testnet_checkbox.setEnabled(can_modify_settings)
        self.connect_button.setEnabled(can_modify_settings)

    def _init_encryption(self):
        """
        初始化加密密钥
        """
        try:
            # 密钥文件路径
            key_file = os.path.join(os.path.dirname(__file__), "encryption_key.key")
            
            if os.path.exists(key_file):
                # 加载现有密钥
                with open(key_file, "rb") as f:
                    self.encryption_key = f.read()
            else:
                # 生成新密钥
                self.encryption_key = Fernet.generate_key()
                # 保存密钥到文件
                with open(key_file, "wb") as f:
                    f.write(self.encryption_key)
                logger.info("生成新的加密密钥")
            
            # 创建加密器
            self.cipher_suite = Fernet(self.encryption_key)
            logger.info("加密系统初始化成功")
        except Exception as e:
            logger.error(f"初始化加密系统错误: {e}")
            # 如果加密初始化失败，使用未加密方式
            self.cipher_suite = None
    
    def _load_encrypted_api_keys(self):
        """
        加载加密的API密钥
        """
        try:
            # API密钥文件路径
            api_file = os.path.join(os.path.dirname(__file__), "api_keys.json")
            
            if os.path.exists(api_file) and self.cipher_suite:
                with open(api_file, "rb") as f:
                    encrypted_data = f.read()
                
                # 解密数据
                decrypted_data = self.cipher_suite.decrypt(encrypted_data)
                api_data = json.loads(decrypted_data.decode('utf-8'))
                
                # 填充API输入框
                if "api_key" in api_data:
                    self.api_key_input.setText(api_data["api_key"])
                if "api_secret" in api_data:
                    self.secret_input.setText(api_data["api_secret"])
                if "passphrase" in api_data:
                    self.passphrase_input.setText(api_data["passphrase"])
                if "is_test" in api_data:
                    self.testnet_checkbox.setCurrentText("模拟盘" if api_data["is_test"] else "实盘")
                
                logger.info("加载加密的API密钥成功")
        except Exception as e:
            logger.error(f"加载加密API密钥错误: {e}")
    
    def _save_encrypted_api_keys(self, api_key, api_secret, passphrase, is_test):
        """
        保存加密的API密钥
        """
        try:
            if not self.cipher_suite:
                logger.warning("加密系统未初始化，跳过API密钥保存")
                return
            
            # API密钥文件路径
            api_file = os.path.join(os.path.dirname(__file__), "api_keys.json")
            
            # 准备数据
            api_data = {
                "api_key": api_key,
                "api_secret": api_secret,
                "passphrase": passphrase,
                "is_test": is_test,
                "updated_at": datetime.now().isoformat()
            }
            
            # 加密数据
            encrypted_data = self.cipher_suite.encrypt(json.dumps(api_data).encode('utf-8'))
            
            # 保存加密数据
            with open(api_file, "wb") as f:
                f.write(encrypted_data)
            
            logger.info("保存加密的API密钥成功")
        except Exception as e:
            logger.error(f"保存加密API密钥错误: {e}")

    def _load_market_data_cache(self):
        """
        加载本地缓存的市场数据
        """
        # 加载默认交易对的市场数据
        default_symbols = ["BTC-USDT-SWAP", "ETH-USDT-SWAP"]
        for symbol in default_symbols:
            cached_data = data_persistence.load_market_data(symbol)
            if cached_data:
                # 更新市场数据存储
                for data in cached_data:
                    if isinstance(data, dict):
                        last_price = data.get("last", "0")
                        change24h = data.get("change24h", "0")
                        change24h_percent = data.get("change24hPercent", "0")
                        vol24h = data.get("vol24h", "0")
                        
                        self.market_data[symbol] = {
                            "last": last_price,
                            "change24h": change24h,
                            "change24hPercent": change24h_percent,
                            "vol24h": vol24h,
                            "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        }
                logger.info(f"加载缓存的市场数据: {symbol}, {len(cached_data)} 条记录")

    def _load_strategies(self):
        """
        加载现有的策略文件
        """
        import os
        try:
            strategies_dir = "d:/Projects/okx_trading_bot/strategies"
            
            if os.path.exists(strategies_dir):
                # 遍历strategies目录中的所有.py文件
                for file_name in os.listdir(strategies_dir):
                    if file_name.endswith(".py") and file_name != "__init__.py" and file_name != "base_strategy.py":
                        # 提取策略名称（去掉.py后缀）
                        strategy_name = os.path.splitext(file_name)[0]
                        
                        # 检查策略是否已经在列表中
                        existing_strategy = next((s for s in self.strategies if s["name"] == strategy_name), None)
                        if not existing_strategy:
                            # 确定策略类型
                            strategy_type = "自定义策略"
                            if "ma_rsi" in strategy_name.lower():
                                strategy_type = "移动平均线RSI策略"
                            elif "macd_bollinger" in strategy_name.lower():
                                strategy_type = "MACD布林带策略"
                            elif "passivbot" in strategy_name.lower():
                                strategy_type = "PassivBot策略"
                            elif "dynamic" in strategy_name.lower():
                                strategy_type = "动态策略"
                            elif "combined" in strategy_name.lower():
                                strategy_type = "组合策略"
                            elif "arbitrage" in strategy_name.lower():
                                strategy_type = "套利策略"
                            elif "machine_learning" in strategy_name.lower():
                                strategy_type = "机器学习策略"
                            elif "nuclear_dynamics" in strategy_name.lower():
                                strategy_type = "原子核互反动力策略"
                            
                            # 添加策略到列表
                            self.strategies.append({
                                "name": strategy_name,
                                "type": strategy_type,
                                "status": "已停止"
                            })
                            print(f"加载策略: {strategy_name} (类型: {strategy_type})")
                            logger.info(f"加载策略: {strategy_name} (类型: {strategy_type})")
            else:
                print(f"策略目录不存在: {strategies_dir}")
                logger.warning(f"策略目录不存在: {strategies_dir}")
                # 创建策略目录
                os.makedirs(strategies_dir, exist_ok=True)
                print(f"创建策略目录: {strategies_dir}")
                logger.info(f"创建策略目录: {strategies_dir}")
        except Exception as e:
            print(f"加载策略错误: {e}")
            logger.error(f"加载策略错误: {e}")

    def init_ui(self):
        """
        初始化用户界面
        """
        self.setWindowTitle("OKX 交易机器人")
        self.setGeometry(100, 100, 1400, 900)
        
        # 设置全局样式
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f0f2f5;
            }
            QGroupBox {
                border: 1px solid #e0e0e0;
                border-radius: 8px;
                margin-top: 10px;
                background-color: white;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 10px 0 10px;
                background-color: white;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton {
                background-color: #1976D2;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
                min-width: 90px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #1565C0;
            }
            QPushButton:disabled {
                background-color: #BDBDBD;
            }
            QLineEdit {
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                padding: 8px;
                font-size: 14px;
            }
            QLineEdit:focus {
                border: 1px solid #1976D2;
            }
            QComboBox {
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                padding: 8px;
                font-size: 14px;
            }
            QTableWidget {
                border: 1px solid #e0e0e0;
                border-radius: 6px;
                alternate-background-color: #f9f9f9;
            }
            QTableWidget::item:selected {
                background-color: #E3F2FD;
            }
            QStatusBar {
                background-color: white;
                border-top: 1px solid #e0e0e0;
                font-size: 12px;
            }
            QLabel {
                font-size: 14px;
            }
            QTabBar::tab {
                background-color: #f5f5f5;
                padding: 12px 20px;
                margin-right: 2px;
                border-radius: 6px 6px 0 0;
                font-size: 14px;
            }
            QTabBar::tab:selected {
                background-color: white;
                border-bottom: 2px solid #1976D2;
            }
        """)

        # 中心部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)

        # 主布局
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)

        # 顶部连接状态栏
        status_group = QGroupBox("连接管理")
        status_layout = QVBoxLayout()
        status_layout.setContentsMargins(15, 15, 15, 15)
        
        # 连接状态和API配置
        api_config_layout = QHBoxLayout()
        api_config_layout.setSpacing(15)
        
        # 连接状态
        status_left = QHBoxLayout()
        status_left.addWidget(QLabel("连接状态:"))
        self.connection_status = QLabel("未连接")
        self.connection_status.setStyleSheet("color: red; font-weight: bold; font-size: 14px;")
        status_left.addWidget(self.connection_status)
        status_left.addStretch(1)
        
        # API配置
        api_inputs_layout = QHBoxLayout()
        api_inputs_layout.setSpacing(10)
        
        api_inputs_layout.addWidget(QLabel("API Key:"))
        self.api_key_input = QLineEdit()
        self.api_key_input.setPlaceholderText("API Key")
        self.api_key_input.setMinimumWidth(220)
        api_inputs_layout.addWidget(self.api_key_input)
        
        api_inputs_layout.addWidget(QLabel("Secret:"))
        self.secret_input = QLineEdit()
        self.secret_input.setPlaceholderText("API Secret")
        self.secret_input.setEchoMode(QLineEdit.Password)
        self.secret_input.setMinimumWidth(220)
        api_inputs_layout.addWidget(self.secret_input)
        
        api_inputs_layout.addWidget(QLabel("Passphrase:"))
        self.passphrase_input = QLineEdit()
        self.passphrase_input.setPlaceholderText("Passphrase")
        self.passphrase_input.setMinimumWidth(180)
        api_inputs_layout.addWidget(self.passphrase_input)
        
        api_inputs_layout.addWidget(QLabel("环境:"))
        self.testnet_checkbox = QComboBox()
        self.testnet_checkbox.addItems(["模拟盘", "实盘"])
        self.testnet_checkbox.setMinimumWidth(100)
        api_inputs_layout.addWidget(self.testnet_checkbox)
        
        self.connect_button = QPushButton("连接")
        self.connect_button.clicked.connect(self.toggle_connection)
        api_inputs_layout.addWidget(self.connect_button)
        
        api_config_layout.addLayout(status_left)
        api_config_layout.addLayout(api_inputs_layout)
        status_layout.addLayout(api_config_layout)
        status_group.setLayout(status_layout)
        main_layout.addWidget(status_group)

        # 仪表盘概览
        dashboard_group = QGroupBox("仪表盘概览")
        self.dashboard_layout = QHBoxLayout()
        self.dashboard_layout.setContentsMargins(15, 15, 15, 15)
        self.dashboard_layout.setSpacing(15)
        
        # 市场概览卡片
        self.market_overview = DraggableDashboardCard("市场概览")
        market_overview_layout = QVBoxLayout()
        market_overview_layout.setContentsMargins(20, 20, 20, 20)
        market_overview_layout.setSpacing(10)
        
        # 市场概览内容
        market_content = QVBoxLayout()
        market_content.setSpacing(8)
        self.market_price_label = QLabel("BTC/USDT: $45,000")
        self.market_price_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        self.market_change_label = QLabel("24h: +2.5%")
        self.market_change_label.setStyleSheet("font-size: 16px; color: green;")
        market_content.addWidget(self.market_price_label)
        market_content.addWidget(self.market_change_label)
        market_overview_layout.addLayout(market_content)
        self.market_overview.setLayout(market_overview_layout)
        
        # 账户概览卡片
        self.account_overview = DraggableDashboardCard("账户概览")
        account_overview_layout = QVBoxLayout()
        account_overview_layout.setContentsMargins(20, 20, 20, 20)
        account_overview_layout.setSpacing(10)
        
        # 账户概览内容
        account_content = QVBoxLayout()
        account_content.setSpacing(8)
        self.account_balance_label = QLabel("总余额: $10,000")
        self.account_balance_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        self.account_pnl_label = QLabel("今日盈亏: +$150")
        self.account_pnl_label.setStyleSheet("font-size: 16px; color: green;")
        account_content.addWidget(self.account_balance_label)
        account_content.addWidget(self.account_pnl_label)
        account_overview_layout.addLayout(account_content)
        self.account_overview.setLayout(account_overview_layout)
        
        # 策略概览卡片
        self.strategy_overview = DraggableDashboardCard("策略概览")
        strategy_overview_layout = QVBoxLayout()
        strategy_overview_layout.setContentsMargins(20, 20, 20, 20)
        strategy_overview_layout.setSpacing(10)
        
        # 策略概览内容
        strategy_content = QVBoxLayout()
        strategy_content.setSpacing(8)
        self.strategy_status_label = QLabel("运行中: 2个策略")
        self.strategy_status_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        self.strategy_profit_label = QLabel("总利润: +$500")
        self.strategy_profit_label.setStyleSheet("font-size: 16px; color: green;")
        strategy_content.addWidget(self.strategy_status_label)
        strategy_content.addWidget(self.strategy_profit_label)
        strategy_overview_layout.addLayout(strategy_content)
        self.strategy_overview.setLayout(strategy_overview_layout)
        
        # 交易概览卡片
        self.trade_overview = DraggableDashboardCard("交易概览")
        trade_overview_layout = QVBoxLayout()
        trade_overview_layout.setContentsMargins(20, 20, 20, 20)
        trade_overview_layout.setSpacing(10)
        
        # 交易概览内容
        trade_content = QVBoxLayout()
        trade_content.setSpacing(8)
        self.trade_count_label = QLabel("今日交易: 12笔")
        self.trade_count_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        self.trade_win_rate_label = QLabel("胜率: 75%")
        self.trade_win_rate_label.setStyleSheet("font-size: 16px; color: green;")
        trade_content.addWidget(self.trade_count_label)
        trade_content.addWidget(self.trade_win_rate_label)
        trade_overview_layout.addLayout(trade_content)
        self.trade_overview.setLayout(trade_overview_layout)
        
        # 风险指标卡片
        self.risk_overview = DraggableDashboardCard("风险指标")
        risk_overview_layout = QVBoxLayout()
        risk_overview_layout.setContentsMargins(20, 20, 20, 20)
        risk_overview_layout.setSpacing(10)
        
        # 风险指标内容
        risk_content = QVBoxLayout()
        risk_content.setSpacing(8)
        self.risk_level_label = QLabel("风险等级: 中等")
        self.risk_level_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        self.max_drawdown_label = QLabel("最大回撤: -3.2%")
        self.max_drawdown_label.setStyleSheet("font-size: 16px; color: red;")
        risk_content.addWidget(self.risk_level_label)
        risk_content.addWidget(self.max_drawdown_label)
        risk_overview_layout.addLayout(risk_content)
        self.risk_overview.setLayout(risk_overview_layout)
        
        # 持仓概览卡片
        self.position_overview = DraggableDashboardCard("持仓概览")
        position_overview_layout = QVBoxLayout()
        position_overview_layout.setContentsMargins(20, 20, 20, 20)
        position_overview_layout.setSpacing(10)
        
        # 持仓概览内容
        position_content = QVBoxLayout()
        position_content.setSpacing(8)
        self.position_count_label = QLabel("持仓数量: 3个")
        self.position_count_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        self.position_value_label = QLabel("持仓价值: $5,000")
        self.position_value_label.setStyleSheet("font-size: 16px; color: green;")
        position_content.addWidget(self.position_count_label)
        position_content.addWidget(self.position_value_label)
        position_overview_layout.addLayout(position_content)
        self.position_overview.setLayout(position_overview_layout)
        
        # 保存仪表盘卡片引用
        self.dashboard_cards = [
            self.market_overview,
            self.account_overview,
            self.strategy_overview,
            self.trade_overview,
            self.risk_overview,
            self.position_overview
        ]
        
        # 添加卡片到仪表盘
        self.dashboard_layout.addWidget(self.market_overview, 1)
        self.dashboard_layout.addWidget(self.account_overview, 1)
        self.dashboard_layout.addWidget(self.strategy_overview, 1)
        self.dashboard_layout.addWidget(self.trade_overview, 1)
        self.dashboard_layout.addWidget(self.risk_overview, 1)
        self.dashboard_layout.addWidget(self.position_overview, 1)
        dashboard_group.setLayout(self.dashboard_layout)
        main_layout.addWidget(dashboard_group)

        # 标签页 - 简化为主要功能
        self.tab_widget = AnimatedTabWidget()
        self.tab_widget.setTabPosition(QTabWidget.North)
        self.tab_widget.setTabShape(QTabWidget.Rounded)
        self.tab_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        main_layout.addWidget(self.tab_widget)
        # 设置标签页为可拉伸
        main_layout.setStretch(0, 0)  # 连接管理部分不拉伸
        main_layout.setStretch(1, 0)  # 仪表盘部分不拉伸
        main_layout.setStretch(2, 1)  # 标签页部分拉伸

        # 市场数据标签页
        self.market_tab = QWidget()
        self.tab_widget.addTab(self.market_tab, "市场数据")
        self.init_market_tab()

        # 交易与订单标签页
        self.trade_tab = QWidget()
        self.tab_widget.addTab(self.trade_tab, "交易与订单")
        self.init_trade_tab()

        # 策略管理标签页
        self.strategy_tab = QWidget()
        self.tab_widget.addTab(self.strategy_tab, "策略管理")
        self.init_strategy_tab()

        # 监控与分析标签页
        self.monitor_tab = QWidget()
        self.tab_widget.addTab(self.monitor_tab, "监控与分析")
        self.init_monitor_tab()

        # 状态栏
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.statusBar.showMessage("就绪")

        # 添加帮助菜单项
        self.init_help_menu()
        
        # 更新UI权限
        self.update_ui_permissions()
        
        # 更新策略列表
        self.update_strategy_list()
        
    def resizeEvent(self, event):
        """
        窗口大小改变事件处理
        """
        super().resizeEvent(event)
        
        # 获取窗口大小
        width = self.width()
        height = self.height()
        
        # 根据窗口宽度调整仪表盘卡片布局
        if hasattr(self, 'dashboard_layout') and hasattr(self, 'dashboard_cards'):
            # 根据窗口宽度决定卡片排列方式
            if width < 1000 and isinstance(self.dashboard_layout, QHBoxLayout):
                # 小屏幕：切换到网格布局
                # 获取父组件
                parent = self.dashboard_cards[0].parent() if self.dashboard_cards else None
                if parent:
                    # 创建网格布局
                    grid_layout = QGridLayout()
                    grid_layout.setContentsMargins(15, 15, 15, 15)
                    grid_layout.setSpacing(15)
                    
                    # 添加卡片到网格布局
                    for i, card in enumerate(self.dashboard_cards):
                        row = i // 2
                        col = i % 2
                        grid_layout.addWidget(card, row, col)
                    
                    # 设置新布局
                    parent.setLayout(grid_layout)
                    self.dashboard_layout = grid_layout
                    logger.debug("切换到网格布局")
            elif width >= 1000 and isinstance(self.dashboard_layout, QGridLayout):
                # 大屏幕：切换到水平布局
                # 获取父组件
                parent = self.dashboard_cards[0].parent() if self.dashboard_cards else None
                if parent:
                    # 创建水平布局
                    h_layout = QHBoxLayout()
                    h_layout.setContentsMargins(15, 15, 15, 15)
                    h_layout.setSpacing(15)
                    
                    # 添加卡片到水平布局
                    for card in self.dashboard_cards:
                        h_layout.addWidget(card, 1)
                    
                    # 设置新布局
                    parent.setLayout(h_layout)
                    self.dashboard_layout = h_layout
                    logger.debug("切换到水平布局")
        
        # 根据窗口大小调整字体大小
        if width < 1200:
            # 小屏幕：减小字体
            self.setStyleSheet(self.styleSheet() + """
                QLabel {
                    font-size: 12px;
                }
                QPushButton {
                    font-size: 12px;
                }
            """)
        
        logger.debug(f"窗口大小已调整: {width}x{height}")

    def init_trade_tab(self):
        """
        初始化交易与订单标签页
        """
        trade_layout = QVBoxLayout(self.trade_tab)
        trade_layout.setContentsMargins(15, 15, 15, 15)
        trade_layout.setSpacing(15)
        
        # 交易操作标签页
        trade_tabs = QTabWidget()
        trade_tabs.setTabPosition(QTabWidget.North)
        trade_tabs.setTabShape(QTabWidget.Rounded)
        
        # 普通交易标签
        normal_trade_tab = QWidget()
        normal_trade_layout = QVBoxLayout(normal_trade_tab)
        normal_trade_layout.setContentsMargins(15, 15, 15, 15)
        normal_trade_layout.setSpacing(15)
        
        # 顶部布局：交易操作
        trade_ops_group = QGroupBox("交易操作")
        trade_ops_layout = QGridLayout()
        trade_ops_layout.setContentsMargins(15, 15, 15, 15)
        trade_ops_layout.setSpacing(10)
        
        # 交易对选择
        trade_ops_layout.addWidget(QLabel("交易对:"), 0, 0)
        self.trade_pair_combo = QComboBox()
        self.trade_pair_combo.addItems(["BTC-USDT-SWAP", "ETH-USDT-SWAP", "BNB-USDT-SWAP"])
        self.trade_pair_combo.setMinimumWidth(150)
        trade_ops_layout.addWidget(self.trade_pair_combo, 0, 1)
        
        # 交易方向
        trade_ops_layout.addWidget(QLabel("方向:"), 0, 2)
        self.trade_side_combo = QComboBox()
        self.trade_side_combo.addItems(["买入", "卖出"])
        self.trade_side_combo.setMinimumWidth(80)
        trade_ops_layout.addWidget(self.trade_side_combo, 0, 3)
        
        # 交易类型
        trade_ops_layout.addWidget(QLabel("类型:"), 0, 4)
        self.trade_type_combo = QComboBox()
        self.trade_type_combo.addItems(["市价", "限价"])
        self.trade_type_combo.setMinimumWidth(80)
        trade_ops_layout.addWidget(self.trade_type_combo, 0, 5)
        
        # 价格
        trade_ops_layout.addWidget(QLabel("价格:"), 1, 0)
        self.trade_price_input = QLineEdit()
        self.trade_price_input.setPlaceholderText("价格")
        self.trade_price_input.setMinimumWidth(120)
        trade_ops_layout.addWidget(self.trade_price_input, 1, 1, 1, 2)
        
        # 数量
        trade_ops_layout.addWidget(QLabel("数量:"), 1, 3)
        self.trade_amount_input = QLineEdit()
        self.trade_amount_input.setPlaceholderText("数量")
        self.trade_amount_input.setMinimumWidth(120)
        trade_ops_layout.addWidget(self.trade_amount_input, 1, 4, 1, 2)
        
        # 执行按钮
        self.execute_trade_button = QPushButton("执行交易")
        self.execute_trade_button.clicked.connect(self.execute_trade)
        self.execute_trade_button.setMinimumWidth(120)
        trade_ops_layout.addWidget(self.execute_trade_button, 2, 0, 1, 6)
        
        trade_ops_group.setLayout(trade_ops_layout)
        normal_trade_layout.addWidget(trade_ops_group)
        
        # 批量交易标签
        batch_trade_tab = QWidget()
        batch_trade_layout = QVBoxLayout(batch_trade_tab)
        batch_trade_layout.setContentsMargins(15, 15, 15, 15)
        batch_trade_layout.setSpacing(15)
        
        # 批量交易操作
        batch_trade_group = QGroupBox("批量交易")
        batch_trade_ops_layout = QVBoxLayout(batch_trade_group)
        batch_trade_ops_layout.setContentsMargins(15, 15, 15, 15)
        
        # 批量交易输入区域
        self.batch_trade_text = QTextEdit()
        self.batch_trade_text.setPlaceholderText("每行输入一个交易指令，格式：交易对,方向,类型,价格,数量\n例如：BTC-USDT-SWAP,买入,限价,40000,0.01")
        self.batch_trade_text.setMinimumHeight(200)
        batch_trade_ops_layout.addWidget(self.batch_trade_text)
        
        # 执行批量交易按钮
        self.execute_batch_trade_button = QPushButton("执行批量交易")
        self.execute_batch_trade_button.clicked.connect(self.execute_batch_trade)
        batch_trade_ops_layout.addWidget(self.execute_batch_trade_button)
        
        batch_trade_group.setLayout(batch_trade_ops_layout)
        batch_trade_layout.addWidget(batch_trade_group)
        
        # 条件单标签
        conditional_trade_tab = QWidget()
        conditional_trade_layout = QVBoxLayout(conditional_trade_tab)
        conditional_trade_layout.setContentsMargins(15, 15, 15, 15)
        conditional_trade_layout.setSpacing(15)
        
        # 条件单操作
        conditional_trade_group = QGroupBox("条件单")
        conditional_trade_ops_layout = QGridLayout(conditional_trade_group)
        conditional_trade_ops_layout.setContentsMargins(15, 15, 15, 15)
        conditional_trade_ops_layout.setSpacing(10)
        
        # 交易对选择
        conditional_trade_ops_layout.addWidget(QLabel("交易对:"), 0, 0)
        self.cond_trade_pair_combo = QComboBox()
        self.cond_trade_pair_combo.addItems(["BTC-USDT-SWAP", "ETH-USDT-SWAP", "BNB-USDT-SWAP"])
        self.cond_trade_pair_combo.setMinimumWidth(150)
        conditional_trade_ops_layout.addWidget(self.cond_trade_pair_combo, 0, 1)
        
        # 交易方向
        conditional_trade_ops_layout.addWidget(QLabel("方向:"), 0, 2)
        self.cond_trade_side_combo = QComboBox()
        self.cond_trade_side_combo.addItems(["买入", "卖出"])
        self.cond_trade_side_combo.setMinimumWidth(80)
        conditional_trade_ops_layout.addWidget(self.cond_trade_side_combo, 0, 3)
        
        # 条件类型
        conditional_trade_ops_layout.addWidget(QLabel("条件类型:"), 1, 0)
        self.cond_type_combo = QComboBox()
        self.cond_type_combo.addItems(["止盈", "止损"])
        self.cond_type_combo.setMinimumWidth(80)
        conditional_trade_ops_layout.addWidget(self.cond_type_combo, 1, 1)
        
        # 触发价格
        conditional_trade_ops_layout.addWidget(QLabel("触发价格:"), 1, 2)
        self.cond_trigger_price_input = QLineEdit()
        self.cond_trigger_price_input.setPlaceholderText("触发价格")
        self.cond_trigger_price_input.setMinimumWidth(120)
        conditional_trade_ops_layout.addWidget(self.cond_trigger_price_input, 1, 3)
        
        # 执行价格
        conditional_trade_ops_layout.addWidget(QLabel("执行价格:"), 2, 0)
        self.cond_execute_price_input = QLineEdit()
        self.cond_execute_price_input.setPlaceholderText("执行价格")
        self.cond_execute_price_input.setMinimumWidth(120)
        conditional_trade_ops_layout.addWidget(self.cond_execute_price_input, 2, 1, 1, 2)
        
        # 数量
        conditional_trade_ops_layout.addWidget(QLabel("数量:"), 2, 3)
        self.cond_amount_input = QLineEdit()
        self.cond_amount_input.setPlaceholderText("数量")
        self.cond_amount_input.setMinimumWidth(120)
        conditional_trade_ops_layout.addWidget(self.cond_amount_input, 2, 4)
        
        # 执行条件单按钮
        self.execute_conditional_trade_button = QPushButton("创建条件单")
        self.execute_conditional_trade_button.clicked.connect(self.execute_conditional_trade)
        self.execute_conditional_trade_button.setMinimumWidth(120)
        conditional_trade_ops_layout.addWidget(self.execute_conditional_trade_button, 3, 0, 1, 5)
        
        conditional_trade_group.setLayout(conditional_trade_ops_layout)
        conditional_trade_layout.addWidget(conditional_trade_group)
        
        # 添加标签页
        trade_tabs.addTab(normal_trade_tab, "普通交易")
        trade_tabs.addTab(batch_trade_tab, "批量交易")
        trade_tabs.addTab(conditional_trade_tab, "条件单")
        
        trade_layout.addWidget(trade_tabs)
        
        # 下方布局：订单和账户信息
        bottom_layout = QHBoxLayout()
        
        # 订单列表
        order_group = QGroupBox("订单列表")
        order_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        order_layout = QVBoxLayout()
        
        self.order_table = QTableWidget()
        self.order_table.setColumnCount(8)
        self.order_table.setHorizontalHeaderLabels(["订单ID", "交易对", "类型", "方向", "价格", "数量", "状态", "时间"])
        self.order_table.horizontalHeader().setStretchLastSection(True)
        self.order_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        order_layout.addWidget(self.order_table)
        order_group.setLayout(order_layout)
        
        # 账户信息
        account_group = QGroupBox("账户信息")
        account_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        account_layout = QVBoxLayout()
        
        self.account_table = QTableWidget()
        self.account_table.setColumnCount(4)
        self.account_table.setHorizontalHeaderLabels(["币种", "可用余额", "冻结余额", "总余额"])
        self.account_table.horizontalHeader().setStretchLastSection(True)
        self.account_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        account_layout.addWidget(self.account_table)
        
        # 资产分布
        asset_distribution_group = QGroupBox("资产分布")
        asset_distribution_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        asset_distribution_layout = QVBoxLayout()
        
        # 创建资产分布图表
        from matplotlib.figure import Figure
        self.asset_figure = Figure(figsize=(5, 3))
        self.asset_canvas = FigureCanvas(self.asset_figure)
        self.asset_canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        asset_distribution_layout.addWidget(self.asset_canvas)
        asset_distribution_group.setLayout(asset_distribution_layout)
        
        account_layout.addWidget(asset_distribution_group)
        account_group.setLayout(account_layout)
        
        bottom_layout.addWidget(order_group, 2)
        bottom_layout.addWidget(account_group, 1)
        
        trade_layout.addLayout(bottom_layout)
        # 设置拉伸比例
        trade_layout.setStretch(0, 0)  # 交易操作部分不拉伸
        trade_layout.setStretch(1, 1)  # 下方布局拉伸
    
    def init_market_tab(self):
        """
        初始化市场数据标签页
        """
        market_layout = QVBoxLayout(self.market_tab)
        market_layout.setContentsMargins(15, 15, 15, 15)
        market_layout.setSpacing(15)
        
        # 顶部控制栏
        control_layout = QHBoxLayout()
        control_layout.setSpacing(10)
        
        # 产品搜索
        self.market_search = QLineEdit()
        self.market_search.setPlaceholderText("搜索产品...")
        self.market_search.setMinimumWidth(200)
        control_layout.addWidget(QLabel("搜索:"))
        control_layout.addWidget(self.market_search)
        
        # 刷新按钮
        refresh_button = QPushButton("刷新")
        refresh_button.clicked.connect(self.update_market_table)
        control_layout.addWidget(refresh_button)
        
        # 显示数量
        self.show_count = QComboBox()
        self.show_count.addItems(["10", "20", "50", "100"])
        self.show_count.setCurrentText("20")
        control_layout.addWidget(QLabel("显示数量:"))
        control_layout.addWidget(self.show_count)
        
        # 时间周期选择
        self.timeframe_combo = QComboBox()
        self.timeframe_combo.addItems(["1分钟", "5分钟", "15分钟", "1小时", "4小时", "1天"])
        self.timeframe_combo.setCurrentText("15分钟")
        control_layout.addWidget(QLabel("时间周期:"))
        control_layout.addWidget(self.timeframe_combo)
        
        # 导出按钮
        export_csv_button = QPushButton("导出CSV")
        export_csv_button.clicked.connect(self.export_market_data_csv)
        control_layout.addWidget(export_csv_button)
        
        export_excel_button = QPushButton("导出Excel")
        export_excel_button.clicked.connect(self.export_market_data_excel)
        control_layout.addWidget(export_excel_button)
        
        control_layout.addStretch(1)
        market_layout.addLayout(control_layout)
        
        # 市场数据和图表区域
        market_chart_layout = QHBoxLayout()
        
        # 市场数据表格
        market_group = QGroupBox("市场数据")
        market_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        market_chart_layout.addWidget(market_group, 1)
        
        market_table_layout = QVBoxLayout(market_group)
        market_table_layout.setContentsMargins(15, 15, 15, 15)
        
        self.market_table = QTableWidget()
        self.market_table.setColumnCount(7)
        self.market_table.setHorizontalHeaderLabels(["交易对", "最新价格", "涨跌幅", "成交量", "最高价", "最低价", "技术指标"])
        self.market_table.horizontalHeader().setStretchLastSection(True)
        self.market_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        market_table_layout.addWidget(self.market_table)
        
        # 图表和指标区域
        chart_indicator_layout = QVBoxLayout()
        
        # 价格图表
        chart_group = QGroupBox("价格图表")
        chart_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        chart_layout = QVBoxLayout(chart_group)
        chart_layout.setContentsMargins(15, 15, 15, 15)
        
        # 图表控制栏
        chart_control_layout = QHBoxLayout()
        chart_control_layout.setSpacing(10)
        
        # 指标选择
        self.indicator_combo = QComboBox()
        self.indicator_combo.addItems(["无", "MA", "RSI", "MACD", "BOLL", "KDJ", "CCI", "ATR"])
        self.indicator_combo.setCurrentText("无")
        self.indicator_combo.currentTextChanged.connect(self.update_price_chart)
        chart_control_layout.addWidget(QLabel("指标:"))
        chart_control_layout.addWidget(self.indicator_combo)
        
        # 产品选择
        self.chart_product_combo = QComboBox()
        self.chart_product_combo.addItems(["BTC-USDT-SWAP", "ETH-USDT-SWAP", "BNB-USDT-SWAP"])
        self.chart_product_combo.currentTextChanged.connect(self.update_price_chart)
        chart_control_layout.addWidget(QLabel("产品:"))
        chart_control_layout.addWidget(self.chart_product_combo)
        
        # 时间范围选择
        self.time_range_combo = QComboBox()
        self.time_range_combo.addItems(["1小时", "4小时", "12小时", "24小时", "3天", "7天"])
        self.time_range_combo.setCurrentText("24小时")
        self.time_range_combo.currentTextChanged.connect(self.update_price_chart)
        chart_control_layout.addWidget(QLabel("时间范围:"))
        chart_control_layout.addWidget(self.time_range_combo)
        
        chart_control_layout.addStretch(1)
        chart_layout.addLayout(chart_control_layout)
        
        # 创建价格图表
        self.price_figure = Figure(figsize=(10, 4))
        self.price_canvas = FigureCanvas(self.price_figure)
        self.price_canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        chart_layout.addWidget(self.price_canvas)
        chart_indicator_layout.addWidget(chart_group)
        
        # 技术指标图表
        indicator_group = QGroupBox("技术指标")
        indicator_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        indicator_layout = QVBoxLayout(indicator_group)
        indicator_layout.setContentsMargins(15, 15, 15, 15)
        
        # 创建指标图表
        self.indicator_figure = Figure(figsize=(10, 2))
        self.indicator_canvas = FigureCanvas(self.indicator_figure)
        self.indicator_canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        indicator_layout.addWidget(self.indicator_canvas)
        chart_indicator_layout.addWidget(indicator_group)
        
        market_chart_layout.addLayout(chart_indicator_layout, 2)
        market_layout.addLayout(market_chart_layout)
        
        # 市场分析工具
        analysis_group = QGroupBox("市场分析")
        analysis_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        analysis_layout = QGridLayout(analysis_group)
        analysis_layout.setContentsMargins(15, 15, 15, 15)
        analysis_layout.setSpacing(15)
        
        # 趋势分析
        trend_analysis = QGroupBox("趋势分析")
        trend_layout = QVBoxLayout(trend_analysis)
        trend_layout.setContentsMargins(10, 10, 10, 10)
        
        self.trend_label = QLabel("当前趋势: 横盘")
        self.trend_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        trend_layout.addWidget(self.trend_label)
        
        self.support_resistance_label = QLabel("支撑位: $38,000 | 阻力位: $42,000")
        trend_layout.addWidget(self.support_resistance_label)
        
        # 趋势强度
        self.trend_strength_label = QLabel("趋势强度: 0")
        trend_layout.addWidget(self.trend_strength_label)
        
        analysis_layout.addWidget(trend_analysis, 0, 0)
        
        # 市场情绪
        sentiment_analysis = QGroupBox("市场情绪")
        sentiment_layout = QVBoxLayout(sentiment_analysis)
        sentiment_layout.setContentsMargins(10, 10, 10, 10)
        
        self.sentiment_label = QLabel("市场情绪: 中性")
        self.sentiment_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        sentiment_layout.addWidget(self.sentiment_label)
        
        # 情绪指标
        self.sentiment_value_label = QLabel("情绪指数: 50/100")
        sentiment_layout.addWidget(self.sentiment_value_label)
        
        # 恐慌与贪婪指数
        self.fear_greed_label = QLabel("恐慌与贪婪指数: 50")
        sentiment_layout.addWidget(self.fear_greed_label)
        
        analysis_layout.addWidget(sentiment_analysis, 0, 1)
        
        # 波动率分析
        volatility_analysis = QGroupBox("波动率分析")
        volatility_layout = QVBoxLayout(volatility_analysis)
        volatility_layout.setContentsMargins(10, 10, 10, 10)
        
        self.volatility_label = QLabel("波动率: 低")
        self.volatility_label.setStyleSheet("font-size: 14px; font-weight: bold;")
        volatility_layout.addWidget(self.volatility_label)
        
        self.volatility_value_label = QLabel("24h波动率: 2.5%")
        volatility_layout.addWidget(self.volatility_value_label)
        
        # ATR值
        self.atr_label = QLabel("ATR: 0.00")
        volatility_layout.addWidget(self.atr_label)
        
        analysis_layout.addWidget(volatility_analysis, 0, 2)
        
        # 量价关系分析
        volume_price_analysis = QGroupBox("量价关系")
        volume_price_layout = QVBoxLayout(volume_price_analysis)
        volume_price_layout.setContentsMargins(10, 10, 10, 10)
        
        self.volume_trend_label = QLabel("量价趋势: 同步")
        volume_price_layout.addWidget(self.volume_trend_label)
        
        self.volume_change_label = QLabel("成交量变化: 0%")
        volume_price_layout.addWidget(self.volume_change_label)
        
        analysis_layout.addWidget(volume_price_analysis, 1, 0)
        
        # 市场广度分析
        market_breadth_analysis = QGroupBox("市场广度")
        market_breadth_layout = QVBoxLayout(market_breadth_analysis)
        market_breadth_layout.setContentsMargins(10, 10, 10, 10)
        
        self.advance_decline_label = QLabel("涨跌比: 1.0")
        market_breadth_layout.addWidget(self.advance_decline_label)
        
        self.market_breadth_label = QLabel("市场广度: 50%")
        market_breadth_layout.addWidget(self.market_breadth_label)
        
        analysis_layout.addWidget(market_breadth_analysis, 1, 1)
        
        # 技术形态识别
        pattern_analysis = QGroupBox("技术形态")
        pattern_layout = QVBoxLayout(pattern_analysis)
        pattern_layout.setContentsMargins(10, 10, 10, 10)
        
        self.pattern_label = QLabel("当前形态: 无")
        pattern_layout.addWidget(self.pattern_label)
        
        self.pattern_strength_label = QLabel("形态强度: 0")
        pattern_layout.addWidget(self.pattern_strength_label)
        
        analysis_layout.addWidget(pattern_analysis, 1, 2)
        
        market_layout.addWidget(analysis_group)
        
        # 设置拉伸比例
        market_layout.setStretch(0, 0)  # 控制栏不拉伸
        market_layout.setStretch(1, 3)  # 市场数据和图表区域拉伸
        market_layout.setStretch(2, 1)  # 市场分析工具拉伸

    def update_price_chart(self, inst_id=None):
        """
        更新价格图表
        """
        try:
            # 获取当前选择的产品
            if inst_id is None:
                inst_id = self.chart_product_combo.currentText()
            
            # 清空图表
            self.price_figure.clear()
            self.indicator_figure.clear()
            
            # 检查是否有价格历史数据
            if inst_id in self.price_history and self.price_history[inst_id]:
                # 获取价格历史数据
                data = self.price_history[inst_id]
                
                # 获取选择的时间范围
                time_range = self.time_range_combo.currentText()
                
                # 根据时间范围过滤数据
                from datetime import datetime, timedelta
                now = datetime.now()
                
                if time_range == "1小时":
                    start_time = now - timedelta(hours=1)
                elif time_range == "4小时":
                    start_time = now - timedelta(hours=4)
                elif time_range == "12小时":
                    start_time = now - timedelta(hours=12)
                elif time_range == "24小时":
                    start_time = now - timedelta(hours=24)
                elif time_range == "3天":
                    start_time = now - timedelta(days=3)
                elif time_range == "7天":
                    start_time = now - timedelta(days=7)
                else:
                    start_time = now - timedelta(hours=24)  # 默认24小时
                
                # 过滤数据
                filtered_data = [(t, p) for t, p in data if t >= start_time]
                
                if not filtered_data:
                    # 如果没有数据，显示提示
                    ax = self.price_figure.add_subplot(111)
                    ax.set_title(f"{inst_id} 价格走势")
                    ax.set_xlabel("时间")
                    ax.set_ylabel("价格 (USDT)")
                    ax.text(0.5, 0.5, "暂无数据", ha='center', va='center', transform=ax.transAxes)
                    
                    # 指标图表也显示提示
                    indicator_ax = self.indicator_figure.add_subplot(111)
                    indicator_ax.set_title("技术指标")
                    indicator_ax.text(0.5, 0.5, "暂无数据", ha='center', va='center', transform=indicator_ax.transAxes)
                else:
                    timestamps = [item[0] for item in filtered_data]
                    prices = [item[1] for item in filtered_data]
                
                # 创建价格图表
                ax = self.price_figure.add_subplot(111)
                ax.plot(timestamps, prices, 'b-', linewidth=2, label="价格")
                
                # 获取选择的指标
                indicator = self.indicator_combo.currentText()
                
                # 绘制技术指标
                if indicator == "MA":
                    # 计算移动平均线
                    import numpy as np
                    prices_np = np.array(prices)
                    ma10 = np.convolve(prices_np, np.ones(10)/10, mode='valid')
                    ma20 = np.convolve(prices_np, np.ones(20)/20, mode='valid')
                    
                    # 绘制移动平均线
                    ax.plot(timestamps[9:], ma10, 'g-', linewidth=1.5, label="MA10")
                    ax.plot(timestamps[19:], ma20, 'r-', linewidth=1.5, label="MA20")
                    ax.legend()
                
                elif indicator == "BOLL":
                    # 计算布林带
                    import numpy as np
                    prices_np = np.array(prices)
                    ma20 = np.convolve(prices_np, np.ones(20)/20, mode='valid')
                    std20 = np.array([np.std(prices_np[i:i+20]) for i in range(len(prices_np)-19)])
                    upper_band = ma20 + 2 * std20
                    lower_band = ma20 - 2 * std20
                    
                    # 绘制布林带
                    ax.plot(timestamps[19:], ma20, 'g-', linewidth=1.5, label="MA20")
                    ax.plot(timestamps[19:], upper_band, 'r--', linewidth=1, label="上轨")
                    ax.plot(timestamps[19:], lower_band, 'g--', linewidth=1, label="下轨")
                    ax.fill_between(timestamps[19:], lower_band, upper_band, alpha=0.1, color='gray')
                    ax.legend()
                
                # 设置图表标题和标签
                ax.set_title(f"{inst_id} 价格走势")
                ax.set_xlabel("时间")
                ax.set_ylabel("价格 (USDT)")
                
                # 设置x轴格式
                import matplotlib.dates as mdates
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
                self.price_figure.autofmt_xdate()
                
                # 添加网格
                ax.grid(True, linestyle='--', alpha=0.7)
                
                # 绘制指标图表
                if indicator in ["RSI", "MACD"]:
                    indicator_ax = self.indicator_figure.add_subplot(111)
                    
                    if indicator == "RSI":
                        # 计算RSI
                        import numpy as np
                        delta = np.diff(prices)
                        gain = (delta > 0) * delta
                        loss = (delta < 0) * -delta
                        
                        # 计算平均增益和平均损失
                        avg_gain = np.convolve(gain, np.ones(14)/14, mode='valid')
                        avg_loss = np.convolve(loss, np.ones(14)/14, mode='valid')
                        
                        # 计算RSI
                        rs = avg_gain / (avg_loss + 1e-10)
                        rsi = 100 - (100 / (1 + rs))
                        
                        # 绘制RSI
                        indicator_ax.plot(timestamps[14:], rsi, 'b-', linewidth=2, label="RSI")
                        indicator_ax.axhline(70, color='r', linestyle='--', alpha=0.7)
                        indicator_ax.axhline(30, color='g', linestyle='--', alpha=0.7)
                        indicator_ax.set_title("RSI 指标")
                        indicator_ax.set_ylabel("RSI")
                        indicator_ax.legend()
                    
                    elif indicator == "MACD":
                        # 计算MACD
                        import numpy as np
                        prices_np = np.array(prices)
                        
                        # 计算EMA
                        def ema(data, period):
                            alpha = 2 / (period + 1)
                            ema_line = np.zeros_like(data)
                            ema_line[period-1] = np.mean(data[:period])
                            for i in range(period, len(data)):
                                ema_line[i] = alpha * data[i] + (1 - alpha) * ema_line[i-1]
                            return ema_line
                        
                        ema12 = ema(prices_np, 12)
                        ema26 = ema(prices_np, 26)
                        macd_line = ema12 - ema26
                        signal_line = ema(macd_line, 9)
                        histogram = macd_line - signal_line
                        
                        # 绘制MACD
                        indicator_ax.plot(timestamps, macd_line, 'b-', linewidth=1.5, label="MACD")
                        indicator_ax.plot(timestamps, signal_line, 'r-', linewidth=1.5, label="Signal")
                        indicator_ax.bar(timestamps, histogram, color='gray', alpha=0.5, label="Histogram")
                        indicator_ax.set_title("MACD 指标")
                        indicator_ax.set_ylabel("MACD")
                        indicator_ax.legend()
                
                elif indicator == "KDJ":
                        # 计算KDJ
                        import numpy as np
                        prices_np = np.array(prices)
                        
                        # 计算最高价和最低价
                        window = 9
                        highest = np.zeros_like(prices_np)
                        lowest = np.zeros_like(prices_np)
                        
                        for i in range(len(prices_np)):
                            start = max(0, i - window + 1)
                            highest[i] = np.max(prices_np[start:i+1])
                            lowest[i] = np.min(prices_np[start:i+1])
                        
                        # 计算RSV
                        rsv = (prices_np - lowest) / (highest - lowest + 1e-10) * 100
                        
                        # 计算K、D、J
                        k = np.zeros_like(rsv)
                        d = np.zeros_like(rsv)
                        j = np.zeros_like(rsv)
                        
                        k[window-1] = 50
                        d[window-1] = 50
                        
                        for i in range(window, len(rsv)):
                            k[i] = 2/3 * k[i-1] + 1/3 * rsv[i]
                            d[i] = 2/3 * d[i-1] + 1/3 * k[i]
                            j[i] = 3 * k[i] - 2 * d[i]
                        
                        # 绘制KDJ
                        indicator_ax.plot(timestamps, k, 'b-', linewidth=1.5, label="K")
                        indicator_ax.plot(timestamps, d, 'r-', linewidth=1.5, label="D")
                        indicator_ax.plot(timestamps, j, 'g-', linewidth=1.5, label="J")
                        indicator_ax.axhline(80, color='r', linestyle='--', alpha=0.7)
                        indicator_ax.axhline(20, color='g', linestyle='--', alpha=0.7)
                        indicator_ax.set_title("KDJ 指标")
                        indicator_ax.set_ylabel("KDJ")
                        indicator_ax.legend()
                
                elif indicator == "CCI":
                        # 计算CCI
                        import numpy as np
                        prices_np = np.array(prices)
                        
                        window = 20
                        cci = np.zeros_like(prices_np)
                        
                        for i in range(window-1, len(prices_np)):
                            # 计算典型价格
                            typical_price = prices_np[i-window+1:i+1].mean()
                            # 计算平均绝对偏差
                            mean_deviation = np.abs(prices_np[i-window+1:i+1] - typical_price).mean()
                            # 计算CCI
                            if mean_deviation > 0:
                                cci[i] = (prices_np[i] - typical_price) / (0.015 * mean_deviation)
                            else:
                                cci[i] = 0
                        
                        # 绘制CCI
                        indicator_ax.plot(timestamps, cci, 'b-', linewidth=1.5, label="CCI")
                        indicator_ax.axhline(100, color='r', linestyle='--', alpha=0.7)
                        indicator_ax.axhline(-100, color='g', linestyle='--', alpha=0.7)
                        indicator_ax.set_title("CCI 指标")
                        indicator_ax.set_ylabel("CCI")
                        indicator_ax.legend()
                
                elif indicator == "ATR":
                        # 计算ATR
                        import numpy as np
                        prices_np = np.array(prices)
                        
                        window = 14
                        atr = np.zeros_like(prices_np)
                        
                        for i in range(1, len(prices_np)):
                            true_range = max(
                                prices_np[i] - prices_np[i-1],
                                abs(prices_np[i] - prices_np[i-1]),
                                abs(prices_np[i-1] - prices_np[i-1])
                            )
                            if i < window:
                                atr[i] = true_range
                            else:
                                atr[i] = (atr[i-1] * (window-1) + true_range) / window
                        
                        # 绘制ATR
                        indicator_ax.plot(timestamps, atr, 'b-', linewidth=1.5, label="ATR")
                        indicator_ax.set_title("ATR 指标")
                        indicator_ax.set_ylabel("ATR")
                        indicator_ax.legend()
                
                # 设置x轴格式
                indicator_ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M:%S'))
                self.indicator_figure.autofmt_xdate()
                indicator_ax.grid(True, linestyle='--', alpha=0.7)
            else:
                # 没有数据时显示提示
                ax = self.price_figure.add_subplot(111)
                ax.set_title(f"{inst_id} 价格走势")
                ax.set_xlabel("时间")
                ax.set_ylabel("价格 (USDT)")
                ax.text(0.5, 0.5, "暂无数据", ha='center', va='center', transform=ax.transAxes)
                
                # 指标图表也显示提示
                indicator_ax = self.indicator_figure.add_subplot(111)
                indicator_ax.set_title("技术指标")
                indicator_ax.text(0.5, 0.5, "暂无数据", ha='center', va='center', transform=indicator_ax.transAxes)
            
            # 重新绘制图表
            self.price_canvas.draw()
            self.indicator_canvas.draw()
        except Exception as e:
            logger.error(f"更新价格图表错误: {e}")

    def init_order_tab(self):
        """
        初始化订单标签页
        """
        layout = QVBoxLayout(self.order_tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # 顶部控制栏
        control_layout = QHBoxLayout()
        
        # 订单状态筛选
        self.order_status_filter = QComboBox()
        self.order_status_filter.addItems(["所有", "待成交", "已成交", "已取消", "部分成交"])
        control_layout.addWidget(QLabel("状态:"))
        control_layout.addWidget(self.order_status_filter)
        
        # 产品筛选
        self.order_product_filter = QLineEdit()
        self.order_product_filter.setPlaceholderText("产品ID")
        self.order_product_filter.setMinimumWidth(150)
        control_layout.addWidget(QLabel("产品:"))
        control_layout.addWidget(self.order_product_filter)
        
        # 刷新按钮
        refresh_button = QPushButton("刷新")
        refresh_button.clicked.connect(self.update_order_table)
        control_layout.addWidget(refresh_button)
        
        # 显示数量
        self.order_show_count = QComboBox()
        self.order_show_count.addItems(["10", "20", "50", "100"])
        self.order_show_count.setCurrentText("20")
        control_layout.addWidget(QLabel("显示数量:"))
        control_layout.addWidget(self.order_show_count)
        
        control_layout.addStretch(1)
        layout.addLayout(control_layout)

        # 订单表格
        self.order_table = QTableWidget()
        self.order_table.setColumnCount(8)
        self.order_table.setHorizontalHeaderLabels(
            ["订单ID", "产品", "类型", "方向", "价格", "数量", "状态", "时间"]
        )
        self.order_table.setSortingEnabled(True)
        self.order_table.setAlternatingRowColors(True)
        self.order_table.setSelectionMode(QTableWidget.SingleSelection)
        self.order_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        # 设置列宽
        self.order_table.setColumnWidth(0, 150)
        self.order_table.setColumnWidth(1, 120)
        self.order_table.setColumnWidth(2, 80)
        self.order_table.setColumnWidth(3, 60)
        self.order_table.setColumnWidth(4, 80)
        self.order_table.setColumnWidth(5, 80)
        self.order_table.setColumnWidth(6, 80)
        self.order_table.setColumnWidth(7, 150)

        layout.addWidget(self.order_table)
        
        # 底部状态栏
        status_layout = QHBoxLayout()
        self.order_status = QLabel("就绪")
        status_layout.addWidget(self.order_status)
        status_layout.addStretch(1)
        layout.addLayout(status_layout)

    def init_account_tab(self):
        """
        初始化账户标签页
        """
        layout = QVBoxLayout(self.account_tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # 账户信息卡片
        account_group = QGroupBox("账户信息")
        account_layout = QHBoxLayout()
        account_layout.setContentsMargins(10, 10, 10, 10)
        
        # 总权益
        equity_group = QGroupBox("总权益")
        equity_layout = QVBoxLayout()
        self.total_equity_label = QLabel("0.00 USDT")
        self.total_equity_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        equity_layout.addWidget(self.total_equity_label)
        equity_group.setLayout(equity_layout)
        account_layout.addWidget(equity_group)
        
        # 保证金比率
        margin_group = QGroupBox("保证金比率")
        margin_layout = QVBoxLayout()
        self.margin_ratio_label = QLabel("0.00%")
        self.margin_ratio_label.setStyleSheet("font-size: 16px;")
        margin_layout.addWidget(self.margin_ratio_label)
        margin_group.setLayout(margin_layout)
        account_layout.addWidget(margin_group)
        
        # 可用余额
        available_group = QGroupBox("可用余额")
        available_layout = QVBoxLayout()
        self.available_balance_label = QLabel("0.00 USDT")
        self.available_balance_label.setStyleSheet("font-size: 16px;")
        available_layout.addWidget(self.available_balance_label)
        available_group.setLayout(available_layout)
        account_layout.addWidget(available_group)
        
        # 总仓位
        position_group = QGroupBox("总仓位")
        position_layout = QVBoxLayout()
        self.total_position_label = QLabel("0.00 USDT")
        self.total_position_label.setStyleSheet("font-size: 16px;")
        position_layout.addWidget(self.total_position_label)
        position_group.setLayout(position_layout)
        account_layout.addWidget(position_group)
        
        account_group.setLayout(account_layout)
        layout.addWidget(account_group)

        # 资产表格
        asset_group = QGroupBox("资产明细")
        asset_layout = QVBoxLayout()
        
        # 资产表格控制栏
        asset_control_layout = QHBoxLayout()
        
        # 资产搜索
        self.asset_search = QLineEdit()
        self.asset_search.setPlaceholderText("搜索币种...")
        self.asset_search.setMinimumWidth(200)
        asset_control_layout.addWidget(QLabel("搜索:"))
        asset_control_layout.addWidget(self.asset_search)
        
        # 刷新按钮
        refresh_button = QPushButton("刷新")
        refresh_button.clicked.connect(self.update_asset_table)
        asset_control_layout.addWidget(refresh_button)
        
        asset_control_layout.addStretch(1)
        asset_layout.addLayout(asset_control_layout)
        
        # 资产表格
        self.asset_table = QTableWidget()
        self.asset_table.setColumnCount(4)
        self.asset_table.setHorizontalHeaderLabels(
            ["币种", "可用余额", "冻结余额", "总余额"]
        )
        self.asset_table.setSortingEnabled(True)
        self.asset_table.setAlternatingRowColors(True)
        self.asset_table.setSelectionMode(QTableWidget.SingleSelection)
        self.asset_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        # 设置列宽
        self.asset_table.setColumnWidth(0, 80)
        self.asset_table.setColumnWidth(1, 120)
        self.asset_table.setColumnWidth(2, 120)
        self.asset_table.setColumnWidth(3, 120)
        
        asset_layout.addWidget(self.asset_table)
        asset_group.setLayout(asset_layout)
        layout.addWidget(asset_group)
        
        # 底部状态栏
        status_layout = QHBoxLayout()
        self.account_status = QLabel("就绪")
        status_layout.addWidget(self.account_status)
        status_layout.addStretch(1)
        layout.addLayout(status_layout)

    def init_subscribe_tab(self):
        """
        初始化订阅管理标签页
        """
        layout = QVBoxLayout(self.subscribe_tab)

        # 订阅控制
        subscribe_control = QHBoxLayout()
        self.inst_id_input = QLineEdit()
        self.inst_id_input.setPlaceholderText("产品ID (如: BTC-USDT-SWAP)")
        self.subscribe_button = QPushButton("订阅")
        self.unsubscribe_button = QPushButton("取消订阅")

        subscribe_control.addWidget(QLabel("产品ID:"))
        subscribe_control.addWidget(self.inst_id_input)
        subscribe_control.addWidget(self.subscribe_button)
        subscribe_control.addWidget(self.unsubscribe_button)

        layout.addLayout(subscribe_control)

        # 订阅列表
        self.subscribe_list = QTableWidget()
        self.subscribe_list.setColumnCount(2)
        self.subscribe_list.setHorizontalHeaderLabels(["产品ID", "状态"])

        layout.addWidget(self.subscribe_list)

        # 连接按钮信号
        self.subscribe_button.clicked.connect(self.subscribe_instrument)
        self.unsubscribe_button.clicked.connect(self.unsubscribe_instrument)

    def init_strategy_tab(self):
        """
        初始化策略管理标签页
        """
        layout = QVBoxLayout(self.strategy_tab)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)

        # 策略控制
        strategy_control = QVBoxLayout()
        
        # 顶部控制栏
        top_control = QHBoxLayout()
        top_control.setSpacing(10)
        
        self.strategy_name_input = QLineEdit()
        self.strategy_name_input.setPlaceholderText("策略名称")
        self.strategy_name_input.setMinimumWidth(150)
        
        self.strategy_type_combo = QComboBox()
        self.strategy_type_combo.addItems(["动态策略", "PassivBot策略", "移动平均线RSI策略", "MACD布林带策略", "原子核互反动力策略", "自定义策略"])
        self.strategy_type_combo.setMinimumWidth(120)
        
        self.create_strategy_button = QPushButton("创建策略")
        self.start_strategy_button = QPushButton("启动策略")
        self.stop_strategy_button = QPushButton("停止策略")
        self.edit_strategy_button = QPushButton("编辑策略")
        self.delete_strategy_button = QPushButton("删除策略")
        self.generate_report_button = QPushButton("生成报告")

        top_control.addWidget(QLabel("策略名称:"))
        top_control.addWidget(self.strategy_name_input)
        top_control.addWidget(QLabel("策略类型:"))
        top_control.addWidget(self.strategy_type_combo)
        top_control.addWidget(self.create_strategy_button)
        top_control.addWidget(self.start_strategy_button)
        top_control.addWidget(self.stop_strategy_button)
        top_control.addWidget(self.edit_strategy_button)
        top_control.addWidget(self.delete_strategy_button)
        top_control.addWidget(self.generate_report_button)
        top_control.addStretch(1)
        
        # 搜索和筛选栏
        filter_control = QHBoxLayout()
        filter_control.setSpacing(10)
        
        self.strategy_search = QLineEdit()
        self.strategy_search.setPlaceholderText("搜索策略...")
        self.strategy_search.setMinimumWidth(200)
        
        self.status_filter = QComboBox()
        self.status_filter.addItems(["所有状态", "运行中", "已停止"])
        self.status_filter.setMinimumWidth(120)
        
        self.type_filter = QComboBox()
        self.type_filter.addItems(["所有类型", "动态策略", "PassivBot策略", "移动平均线RSI策略", "MACD布林带策略", "原子核互反动力策略", "自定义策略"])
        self.type_filter.setMinimumWidth(120)
        
        refresh_button = QPushButton("刷新")
        refresh_button.clicked.connect(self.update_strategy_list)
        
        filter_control.addWidget(QLabel("搜索:"))
        filter_control.addWidget(self.strategy_search)
        filter_control.addWidget(QLabel("状态:"))
        filter_control.addWidget(self.status_filter)
        filter_control.addWidget(QLabel("类型:"))
        filter_control.addWidget(self.type_filter)
        filter_control.addWidget(refresh_button)
        filter_control.addStretch(1)
        
        strategy_control.addLayout(top_control)
        strategy_control.addLayout(filter_control)

        layout.addLayout(strategy_control)

        # 策略列表
        strategy_list_group = QGroupBox("策略列表")
        strategy_list_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        strategy_list_layout = QVBoxLayout(strategy_list_group)
        strategy_list_layout.setContentsMargins(15, 15, 15, 15)
        
        self.strategy_list = QTableWidget()
        self.strategy_list.setColumnCount(5)
        self.strategy_list.setHorizontalHeaderLabels(
            ["策略名称", "策略类型", "状态", "性能", "操作"]
        )
        # 设置列宽
        self.strategy_list.setColumnWidth(0, 150)
        self.strategy_list.setColumnWidth(1, 120)
        self.strategy_list.setColumnWidth(2, 80)
        self.strategy_list.setColumnWidth(3, 120)
        self.strategy_list.setColumnWidth(4, 300)
        self.strategy_list.horizontalHeader().setStretchLastSection(True)
        self.strategy_list.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        strategy_list_layout.addWidget(self.strategy_list)
        layout.addWidget(strategy_list_group)
        
        # 策略性能分析
        performance_group = QGroupBox("策略性能分析")
        performance_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        performance_layout = QVBoxLayout(performance_group)
        performance_layout.setContentsMargins(15, 15, 15, 15)
        
        # 创建性能分析图表
        self.performance_figure = Figure(figsize=(10, 4))
        self.performance_canvas = FigureCanvas(self.performance_figure)
        self.performance_canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        performance_layout.addWidget(self.performance_canvas)
        layout.addWidget(performance_group)
        
        # 设置拉伸比例
        layout.setStretch(0, 0)  # 控制栏不拉伸
        layout.setStretch(1, 2)  # 策略列表拉伸
        layout.setStretch(2, 1)  # 性能分析拉伸

        # 连接按钮信号
        self.create_strategy_button.clicked.connect(self.create_strategy)
        self.start_strategy_button.clicked.connect(self.start_strategy)
        self.stop_strategy_button.clicked.connect(self.stop_strategy)
        self.edit_strategy_button.clicked.connect(self.edit_strategy)
        self.delete_strategy_button.clicked.connect(self.delete_strategy)
        self.generate_report_button.clicked.connect(self.generate_strategy_report)

        # 初始化策略列表
        self.update_strategy_list()
        # 初始化性能图表
        self.update_performance_chart()

    def toggle_connection(self):
        """
        切换 WebSocket 连接状态
        """
        if self.ws_client and self.ws_client.is_connected():
            # 断开连接
            self._run_async_task(self.disconnect_ws())
        else:
            # 连接
            api_key = self.api_key_input.text()
            api_secret = self.secret_input.text()
            passphrase = self.passphrase_input.text()
            is_test = self.testnet_checkbox.currentText() == "模拟盘"

            self._run_async_task(
                self.connect_ws(api_key, api_secret, passphrase, is_test)
            )

    def _run_async_task(self, coro):
        """
        安全地运行异步任务
        """

        def run_coro():
            try:
                asyncio.run(coro)
            except Exception as e:
                logger.error(f"运行异步任务错误: {e}")
                # 显示错误信息
                self.statusBar.showMessage(f"任务执行错误: {str(e)}")

        # 使用线程池运行任务
        try:
            self._thread_pool.submit(run_coro)
        except Exception as e:
            logger.error(f"提交任务到线程池错误: {e}")
            self.statusBar.showMessage(f"任务提交失败: {str(e)}")

    async def connect_ws(self, api_key, api_secret, passphrase, is_test):
        """
        连接 WebSocket
        """
        try:
            self.ws_client = OKXWebSocketClient(
                api_key=api_key,
                api_secret=api_secret,
                passphrase=passphrase,
                is_test=is_test,
            )

            self.statusBar.showMessage("正在连接 WebSocket...")

            # 连接
            success = await self.ws_client.connect()

            if success:
                self.update_connection_status.emit(True, "已连接")
                self.statusBar.showMessage("WebSocket 连接成功")
                self.connect_button.setText("断开")
                
                # 保存加密的API密钥
                self._save_encrypted_api_keys(api_key, api_secret, passphrase, is_test)

                # 启动市场数据智能体
                if not self.market_data_agent:
                    market_data_config = AgentConfig(name="MarketDataAgent", description="市场数据智能体")
                    self.market_data_agent = MarketDataAgent(
                        market_data_config,
                        api_key=api_key,
                        api_secret=api_secret,
                        passphrase=passphrase,
                        is_test=is_test
                    )
                    await self.market_data_agent.start()
                    logger.info("市场数据智能体已启动")

                # 启动自动重连机制
                self.start_reconnect_timer()

            else:
                self.update_connection_status.emit(False, "连接失败")
                self.statusBar.showMessage("WebSocket 连接失败")
                # 尝试重连
                self.schedule_reconnect(api_key, api_secret, passphrase, is_test)

        except Exception as e:
            logger.error(f"连接 WebSocket 错误: {e}")
            self.update_connection_status.emit(False, f"连接错误: {str(e)}")
            self.statusBar.showMessage(f"连接错误: {str(e)}")
            # 尝试重连
            self.schedule_reconnect(api_key, api_secret, passphrase, is_test)

    async def disconnect_ws(self):
        """
        断开 WebSocket 连接
        """
        try:
            # 停止自动重连机制
            self.stop_reconnect_timer()
            
            if self.ws_client:
                await self.ws_client.close()
                
            # 停止市场数据智能体
            if self.market_data_agent:
                await self.market_data_agent.stop()
                self.market_data_agent = None
                logger.info("市场数据智能体已停止")
                
            self.update_connection_status.emit(False, "已断开")
            self.statusBar.showMessage("WebSocket 已断开")
            self.connect_button.setText("连接")
        except Exception as e:
            logger.error(f"断开 WebSocket 错误: {e}")
            self.statusBar.showMessage(f"断开连接错误: {str(e)}")

    def start_reconnect_timer(self):
        """
        启动自动重连定时器
        """
        # 停止之前的定时器
        self.stop_reconnect_timer()
        
        # 创建新的定时器，每30秒检查一次连接状态
        self.reconnect_timer = QTimer(self)
        self.reconnect_timer.timeout.connect(self.check_connection)
        self.reconnect_timer.start(30000)  # 30秒检查一次
        logger.info("自动重连机制已启动")

    def stop_reconnect_timer(self):
        """
        停止自动重连定时器
        """
        if hasattr(self, 'reconnect_timer') and self.reconnect_timer:
            self.reconnect_timer.stop()
            delattr(self, 'reconnect_timer')
            logger.info("自动重连机制已停止")

    def check_connection(self):
        """
        检查连接状态
        """
        try:
            if self.ws_client and not self.ws_client.is_connected():
                logger.warning("WebSocket 连接已断开，尝试重连...")
                self.statusBar.showMessage("连接已断开，尝试重连...")
                
                # 尝试重连
                api_key = self.api_key_input.text()
                api_secret = self.secret_input.text()
                passphrase = self.passphrase_input.text()
                is_test = self.testnet_checkbox.currentText() == "模拟盘"
                
                if api_key and api_secret and passphrase:
                    self._run_async_task(self.connect_ws(api_key, api_secret, passphrase, is_test))
                else:
                    logger.error("API 配置不完整，无法自动重连")
                    self.statusBar.showMessage("API 配置不完整，无法自动重连")
        except Exception as e:
            logger.error(f"检查连接状态错误: {e}")

    def schedule_reconnect(self, api_key, api_secret, passphrase, is_test):
        """
        安排重连
        """
        # 3秒后尝试重连
        QTimer.singleShot(3000, lambda: self._run_async_task(self.connect_ws(api_key, api_secret, passphrase, is_test)))
        logger.info("已安排 3 秒后尝试重连")

    def start_health_check_timer(self):
        """
        启动系统健康检查定时器
        """
        # 创建定时器，每60秒执行一次健康检查
        self.health_check_timer = QTimer(self)
        self.health_check_timer.timeout.connect(self.perform_health_check)
        self.health_check_timer.start(60000)  # 60秒检查一次
        logger.info("系统健康检查定时器已启动")

    def perform_health_check(self):
        """
        执行系统健康检查
        """
        try:
            # 更新系统资源使用情况
            self.update_system_resources()
            
            # 更新系统健康状态
            self.update_system_health()
            
            # 更新连接状态
            self.update_connection_status_display()
            
            # 更新策略状态
            self.update_strategy_status()
            
            logger.info("系统健康检查完成")
        except Exception as e:
            logger.error(f"执行健康检查错误: {e}")

    def update_system_resources(self):
        """
        更新系统资源使用情况
        """
        try:
            # 尝试导入psutil模块获取系统资源使用情况
            try:
                import psutil
                # 获取CPU使用率
                cpu_usage = psutil.cpu_percent(interval=0.1)
                self.cpu_usage_label.setText(f"{cpu_usage:.1f}%")
                
                # 获取内存使用率
                memory = psutil.virtual_memory()
                memory_usage = memory.percent
                self.memory_usage_label.setText(f"{memory_usage:.1f}%")
            except ImportError:
                # 如果psutil模块不可用，显示默认值
                logger.warning("psutil模块不可用，无法获取系统资源使用情况")
        except Exception as e:
            logger.error(f"更新系统资源使用情况错误: {e}")

    def update_system_health(self):
        """
        更新系统健康状态
        """
        try:
            # 检查各项系统状态
            is_healthy = True
            issues = []
            
            # 检查WebSocket连接状态
            if not (self.ws_client and self.ws_client.is_connected()):
                is_healthy = False
                issues.append("WebSocket连接断开")
            
            # 检查市场数据智能体状态
            if not self.market_data_agent:
                is_healthy = False
                issues.append("市场数据智能体未启动")
            
            # 检查策略状态
            running_strategies = sum(1 for s in self.strategies if s.get('status') == '运行中')
            if running_strategies > 0:
                # 检查策略线程状态
                for strategy_name, thread in self.strategy_threads.items():
                    if not thread.is_alive():
                        is_healthy = False
                        issues.append(f"策略 {strategy_name} 线程已停止")
            
            # 更新系统健康状态标签
            if is_healthy:
                self.system_health_label.setText("正常")
                self.system_health_label.setStyleSheet("font-weight: bold; color: green;")
            else:
                self.system_health_label.setText(f"异常 ({len(issues)}个问题)")
                self.system_health_label.setStyleSheet("font-weight: bold; color: red;")
                # 记录问题
                for issue in issues:
                    logger.warning(f"系统健康检查发现问题: {issue}")
        except Exception as e:
            logger.error(f"更新系统健康状态错误: {e}")

    def update_connection_status_display(self):
        """
        更新连接状态显示
        """
        try:
            # 更新WebSocket连接状态
            if self.ws_client and self.ws_client.is_connected():
                self.ws_status_label.setText("已连接")
                self.ws_status_label.setStyleSheet("font-weight: bold; color: green;")
            else:
                self.ws_status_label.setText("未连接")
                self.ws_status_label.setStyleSheet("font-weight: bold; color: red;")
            
            # 更新市场数据智能体状态
            if self.market_data_agent:
                self.agent_status_label.setText("已启动")
                self.agent_status_label.setStyleSheet("font-weight: bold; color: green;")
            else:
                self.agent_status_label.setText("未启动")
                self.agent_status_label.setStyleSheet("font-weight: bold; color: red;")
        except Exception as e:
            logger.error(f"更新连接状态显示错误: {e}")

    def update_strategy_status(self):
        """
        更新策略状态
        """
        try:
            # 统计运行中的策略数量
            running_strategies = sum(1 for s in self.strategies if s.get('status') == '运行中')
            self.strategy_count_label.setText(f"{running_strategies}个")
        except Exception as e:
            logger.error(f"更新策略状态错误: {e}")

    def subscribe_instrument(self):
        """
        订阅产品
        """
        inst_id = self.inst_id_input.text().strip()
        if not inst_id:
            self.statusBar.showMessage("请输入产品ID")
            return

        if not self.ws_client or not self.ws_client.is_connected():
            self.statusBar.showMessage("请先连接 WebSocket")
            return

        self._run_async_task(self._subscribe_instrument(inst_id))

    async def _subscribe_instrument(self, inst_id):
        """
        异步订阅产品
        """
        try:
            if self.market_data_agent:
                success = await self.market_data_agent.subscribe_instrument(inst_id)
                if success:
                    self.statusBar.showMessage(f"订阅 {inst_id} 成功")
                    # 更新订阅列表
                    self.update_subscribe_list()
                else:
                    self.statusBar.showMessage(f"订阅 {inst_id} 失败")
            else:
                success = await self.ws_client.subscribe("tickers", inst_id)
                if success:
                    self.statusBar.showMessage(f"订阅 {inst_id} 成功")
                    # 更新订阅列表
                    self.update_subscribe_list()
                else:
                    self.statusBar.showMessage(f"订阅 {inst_id} 失败")
        except Exception as e:
            logger.error(f"订阅错误: {e}")
            self.statusBar.showMessage(f"订阅错误: {str(e)}")

    def unsubscribe_instrument(self):
        """
        取消订阅产品
        """
        # 获取选中的行
        selected_rows = self.subscribe_list.selectionModel().selectedRows()
        if not selected_rows:
            self.statusBar.showMessage("请选择要取消订阅的产品")
            return

        if not self.ws_client or not self.ws_client.is_connected():
            self.statusBar.showMessage("请先连接 WebSocket")
            return

        # 取消订阅选中的产品
        for row in selected_rows:
            inst_id = self.subscribe_list.item(row.row(), 0).text()
            self._run_async_task(self._unsubscribe_instrument(inst_id))

    async def _unsubscribe_instrument(self, inst_id):
        """
        异步取消订阅产品
        """
        try:
            if self.market_data_agent:
                success = await self.market_data_agent.unsubscribe_instrument(inst_id)
                if success:
                    self.statusBar.showMessage(f"取消订阅 {inst_id} 成功")
                    # 更新订阅列表
                    self.update_subscribe_list()
                else:
                    self.statusBar.showMessage(f"取消订阅 {inst_id} 失败")
            else:
                # 直接使用WebSocket客户端取消订阅
                success = await self.ws_client.unsubscribe("tickers", inst_id)
                if success:
                    self.statusBar.showMessage(f"取消订阅 {inst_id} 成功")
                    # 更新订阅列表
                    self.update_subscribe_list()
                else:
                    self.statusBar.showMessage(f"取消订阅 {inst_id} 失败")
        except Exception as e:
            logger.error(f"取消订阅错误: {e}")
            self.statusBar.showMessage(f"取消订阅错误: {str(e)}")

    def update_subscribe_list(self):
        """
        更新订阅列表
        """
        # 清空表格
        self.subscribe_list.setRowCount(0)

        # 添加默认订阅的产品
        default_products = ["BTC-USDT-SWAP", "ETH-USDT-SWAP"]
        for product in default_products:
            row_position = self.subscribe_list.rowCount()
            self.subscribe_list.insertRow(row_position)
            self.subscribe_list.setItem(row_position, 0, QTableWidgetItem(product))
            self.subscribe_list.setItem(row_position, 1, QTableWidgetItem("已订阅"))

    def create_strategy(self):
        """
        创建策略
        """
        strategy_name = self.strategy_name_input.text().strip()
        if not strategy_name:
            self.statusBar.showMessage("请输入策略名称")
            return

        strategy_type = self.strategy_type_combo.currentText()

        # 检查策略是否已存在
        existing_strategy_names = [s["name"] for s in self.strategies]
        if strategy_name in existing_strategy_names:
            self.statusBar.showMessage(f"策略 {strategy_name} 已存在")
            return

        # 预定义策略模板
        strategy_templates = {
            "移动平均线RSI策略": """import time
import numpy as np
import logging

logger = logging.getLogger("Strategy")
from strategies.base_strategy import BaseStrategy


class {strategy_name}(BaseStrategy):
    '''移动平均线和RSI结合的交易策略'''

    def __init__(self, api_client=None, config=None):
        '''
        初始化移动平均线和RSI策略

        Args:
            api_client (OKXAPIClient, optional): OKX API客户端实例
            config (dict, optional): 策略配置
        '''
        super().__init__(api_client, config)

        # 策略参数
        self.strategy_params = {
            "ma_short": 10,  # 短期移动平均线周期
            "ma_long": 30,   # 长期移动平均线周期
            "rsi_period": 14, # RSI周期
            "rsi_overbought": 70,  # RSI超买阈值
            "rsi_oversold": 30,    # RSI超卖阈值
            "trend_threshold": 0.001,  # 趋势阈值
        }

        # 数据容器
        self.price_history = []
        self.ma_short_history = []
        self.ma_long_history = []
        self.rsi_history = []

        # 更新配置
        if config and "strategy" in config:
            self.strategy_params.update(config["strategy"])

        logger.info("移动平均线和RSI策略初始化完成")

    def calculate_ma(self, prices, period):
        '''计算移动平均线'''
        if len(prices) < period:
            return None
        return np.mean(prices[-period:])

    def calculate_rsi(self, prices, period):
        '''计算RSI指标'''
        if len(prices) < period + 1:
            return None
        
        deltas = np.diff(prices)
        gains = deltas[deltas > 0]
        losses = -deltas[deltas < 0]
        
        avg_gain = np.mean(gains[-period:]) if len(gains) > 0 else 0
        avg_loss = np.mean(losses[-period:]) if len(losses) > 0 else 0
        
        if avg_loss == 0:
            return 100
        
        rs = avg_gain / avg_loss
        rsi = 100 - (100 / (1 + rs))
        return rsi

    def calculate_trend(self):
        '''计算趋势'''
        if len(self.ma_short_history) < 2 or len(self.ma_long_history) < 2:
            return 0
        
        # 短期均线趋势
        ma_short_trend = self.ma_short_history[-1] - self.ma_short_history[-2]
        # 长期均线趋势
        ma_long_trend = self.ma_long_history[-1] - self.ma_long_history[-2]
        # 均线差趋势
        ma_diff = (self.ma_short_history[-1] - self.ma_long_history[-1])
        
        # 综合趋势
        trend = 0
        if ma_short_trend > self.strategy_params["trend_threshold"] and ma_long_trend > 0:
            trend = 1  # 多头趋势
        elif ma_short_trend < -self.strategy_params["trend_threshold"] and ma_long_trend < 0:
            trend = -1  # 空头趋势
        
        return trend

    def _execute_strategy(self, market_data):
        '''执行策略，生成交易信号

        Args:
            market_data (dict): 市场数据

        Returns:
            dict: 交易信号，包含side, price, amount等信息
        '''
        # 保存当前价格到历史数据
        if "price" in market_data:
            self.price_history.append(market_data["price"])
        elif "last" in market_data:
            self.price_history.append(float(market_data["last"]))
        else:
            logger.warning("市场数据中没有价格信息")
            return None

        # 计算移动平均线
        ma_short = self.calculate_ma(self.price_history, self.strategy_params["ma_short"])
        ma_long = self.calculate_ma(self.price_history, self.strategy_params["ma_long"])
        
        if ma_short:
            self.ma_short_history.append(ma_short)
        if ma_long:
            self.ma_long_history.append(ma_long)

        # 计算RSI
        rsi = self.calculate_rsi(self.price_history, self.strategy_params["rsi_period"])
        if rsi:
            self.rsi_history.append(rsi)

        # 获取当前价格
        current_price = self.price_history[-1]

        # 生成交易信号
        side = "neutral"
        signal_strength = 0

        # 趋势判断
        trend = self.calculate_trend()

        # 金叉死叉信号
        if len(self.ma_short_history) > 1 and len(self.ma_long_history) > 1:
            # 金叉：短期均线上穿长期均线
            if self.ma_short_history[-1] > self.ma_long_history[-1] and self.ma_short_history[-2] <= self.ma_long_history[-2]:
                if rsi and rsi < 50:
                    side = "buy"
                    signal_strength = 0.7
            # 死叉：短期均线下穿长期均线
            elif self.ma_short_history[-1] < self.ma_long_history[-1] and self.ma_short_history[-2] >= self.ma_long_history[-2]:
                if rsi and rsi > 50:
                    side = "sell"
                    signal_strength = -0.7

        # RSI超买超卖信号
        if rsi:
            if rsi < self.strategy_params["rsi_oversold"] and trend >= 0:
                side = "buy"
                signal_strength = 0.8
            elif rsi > self.strategy_params["rsi_overbought"] and trend <= 0:
                side = "sell"
                signal_strength = -0.8

        # 构建交易信号
        signal = {
            "strategy": self.name,
            "side": side,
            "price": current_price,
            "signal_strength": signal_strength,
            "timestamp": market_data.get("timestamp", time.time()),
            "inst_id": market_data.get("inst_id", "BTC-USDT-SWAP"),
            "indicators": {
                "ma_short": ma_short,
                "ma_long": ma_long,
                "rsi": rsi,
                "trend": trend
            }
        }

        logger.info(f"策略信号生成: {signal}")
        return signal
""",
            "MACD布林带策略": """import time
import numpy as np
import logging

logger = logging.getLogger("Strategy")
from strategies.base_strategy import BaseStrategy


class {strategy_name}(BaseStrategy):
    '''MACD和布林带结合的交易策略'''

    def __init__(self, api_client=None, config=None):
        '''
        初始化MACD和布林带策略

        Args:
            api_client (OKXAPIClient, optional): OKX API客户端实例
            config (dict, optional): 策略配置
        '''
        super().__init__(api_client, config)

        # 策略参数
        self.strategy_params = {
            "macd_fast": 12,     # MACD快线周期
            "macd_slow": 26,     # MACD慢线周期
            "macd_signal": 9,    # MACD信号线周期
            "bollinger_period": 20,  # 布林带周期
            "bollinger_std": 2,      # 布林带标准差倍数
            "signal_threshold": 0.001,  # 信号阈值
        }

        # 数据容器
        self.price_history = []
        self.macd_history = []
        self.signal_history = []
        self.histogram_history = []
        self.bollinger_upper = []
        self.bollinger_middle = []
        self.bollinger_lower = []

        # 更新配置
        if config and "strategy" in config:
            self.strategy_params.update(config["strategy"])

        logger.info("MACD和布林带策略初始化完成")

    def calculate_ema(self, prices, period):
        '''计算指数移动平均线'''
        if len(prices) < period:
            return None
        
        ema = []
        multiplier = 2 / (period + 1)
        # 初始EMA为简单移动平均
        initial_ema = np.mean(prices[:period])
        ema.append(initial_ema)
        
        # 计算后续EMA
        for price in prices[period:]:
            current_ema = (price - ema[-1]) * multiplier + ema[-1]
            ema.append(current_ema)
        
        return ema[-1]

    def calculate_macd(self, prices):
        '''计算MACD指标'''
        if len(prices) < max(self.strategy_params["macd_slow"], self.strategy_params["macd_signal"]):
            return None, None, None
        
        # 计算EMA
        ema_fast = self.calculate_ema(prices, self.strategy_params["macd_fast"])
        ema_slow = self.calculate_ema(prices, self.strategy_params["macd_slow"])
        
        if ema_fast is None or ema_slow is None:
            return None, None, None
        
        # 计算MACD线
        macd_line = ema_fast - ema_slow
        
        # 计算信号线
        if len(self.macd_history) >= self.strategy_params["macd_signal"]:
            signal_line = self.calculate_ema(self.macd_history, self.strategy_params["macd_signal"])
        else:
            signal_line = None
        
        # 计算柱状图
        histogram = macd_line - signal_line if signal_line is not None else None
        
        return macd_line, signal_line, histogram

    def calculate_bollinger_bands(self, prices):
        '''计算布林带'''
        if len(prices) < self.strategy_params["bollinger_period"]:
            return None, None, None
        
        # 计算移动平均
        middle_band = np.mean(prices[-self.strategy_params["bollinger_period":])
        # 计算标准差
        std_dev = np.std(prices[-self.strategy_params["bollinger_period":])
        # 计算上下轨
        upper_band = middle_band + (self.strategy_params["bollinger_std"] * std_dev)
        lower_band = middle_band - (self.strategy_params["bollinger_std"] * std_dev)
        
        return upper_band, middle_band, lower_band

    def _execute_strategy(self, market_data):
        '''执行策略，生成交易信号

        Args:
            market_data (dict): 市场数据

        Returns:
            dict: 交易信号，包含side, price, amount等信息
        '''
        # 保存当前价格到历史数据
        if "price" in market_data:
            self.price_history.append(market_data["price"])
        elif "last" in market_data:
            self.price_history.append(float(market_data["last"]))
        else:
            logger.warning("市场数据中没有价格信息")
            return None

        # 计算MACD
        macd_line, signal_line, histogram = self.calculate_macd(self.price_history)
        
        if macd_line:
            self.macd_history.append(macd_line)
        if signal_line:
            self.signal_history.append(signal_line)
        if histogram:
            self.histogram_history.append(histogram)

        # 计算布林带
        upper_band, middle_band, lower_band = self.calculate_bollinger_bands(self.price_history)
        
        if upper_band:
            self.bollinger_upper.append(upper_band)
        if middle_band:
            self.bollinger_middle.append(middle_band)
        if lower_band:
            self.bollinger_lower.append(lower_band)

        # 获取当前价格
        current_price = self.price_history[-1]

        # 生成交易信号
        side = "neutral"
        signal_strength = 0

        # MACD信号
        if len(self.macd_history) > 1 and len(self.signal_history) > 1:
            # MACD金叉：MACD线上穿信号线
            if self.macd_history[-1] > self.signal_history[-1] and self.macd_history[-2] <= self.signal_history[-2]:
                if histogram and histogram > 0:
                    side = "buy"
                    signal_strength = 0.6
            # MACD死叉：MACD线下穿信号线
            elif self.macd_history[-1] < self.signal_history[-1] and self.macd_history[-2] >= self.signal_history[-2]:
                if histogram and histogram < 0:
                    side = "sell"
                    signal_strength = -0.6

        # 布林带信号
        if upper_band and lower_band and middle_band:
            # 价格突破上轨
            if current_price > upper_band:
                side = "sell"
                signal_strength = -0.7
            # 价格突破下轨
            elif current_price < lower_band:
                side = "buy"
                signal_strength = 0.7
            # 价格回归中轨
            elif abs(current_price - middle_band) < self.strategy_params["signal_threshold"]:
                # 根据MACD方向决定
                if len(self.histogram_history) > 0 and self.histogram_history[-1] > 0:
                    side = "buy"
                    signal_strength = 0.5
                elif len(self.histogram_history) > 0 and self.histogram_history[-1] < 0:
                    side = "sell"
                    signal_strength = -0.5

        # 构建交易信号
        signal = {
            "strategy": self.name,
            "side": side,
            "price": current_price,
            "signal_strength": signal_strength,
            "timestamp": market_data.get("timestamp", time.time()),
            "inst_id": market_data.get("inst_id", "BTC-USDT-SWAP"),
            "indicators": {
                "macd_line": macd_line,
                "signal_line": signal_line,
                "histogram": histogram,
                "bollinger_upper": upper_band,
                "bollinger_middle": middle_band,
                "bollinger_lower": lower_band
            }
        }

        logger.info(f"策略信号生成: {signal}")
        return signal
"""
        }

        # 打开代码导入对话框
        from PyQt5.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QTextEdit,
            QPushButton,
            QLabel,
            QHBoxLayout,
        )

        dialog = QDialog(self)
        dialog.setWindowTitle("导入策略代码")
        dialog.setGeometry(200, 200, 800, 600)

        layout = QVBoxLayout(dialog)

        # 提示信息
        info_label = QLabel("请输入或粘贴策略代码:")
        layout.addWidget(info_label)

        # 代码输入区域
        code_editor = QTextEdit()
        
        # 如果是预定义策略类型，加载模板
        if strategy_type in strategy_templates:
            template = strategy_templates[strategy_type].format(strategy_name=strategy_name)
            code_editor.setPlainText(template)
        else:
            code_editor.setPlaceholderText("请输入策略代码...")
            
        layout.addWidget(code_editor)

        # 按钮布局
        button_layout = QHBoxLayout()
        cancel_button = QPushButton("取消")
        create_button = QPushButton("创建")

        button_layout.addStretch(1)
        button_layout.addWidget(cancel_button)
        button_layout.addWidget(create_button)

        layout.addLayout(button_layout)

        # 连接按钮信号
        def on_cancel():
            dialog.reject()

        def on_create():
            code = code_editor.toPlainText().strip()
            if not code:
                self.statusBar.showMessage("请输入策略代码")
                return

            # 保存策略代码到文件
            import os

            strategies_dir = "d:/Projects/okx_trading_bot/strategies"
            strategy_file = os.path.join(strategies_dir, f"{strategy_name}.py")

            try:
                with open(strategy_file, "w", encoding="utf-8") as f:
                    f.write(code)
                logger.info(f"策略文件已创建: {strategy_file}")

                # 添加新策略到列表
                self.strategies.append(
                    {
                        "name": strategy_name,
                        "type": strategy_type,
                        "status": "已停止",
                        "file": f"{strategy_name}.py",
                    }
                )

                self.statusBar.showMessage(
                    f"创建策略 {strategy_name} ({strategy_type}) 成功"
                )

                # 更新策略列表
                self.update_strategy_list()

                dialog.accept()
            except Exception as e:
                logger.error(f"创建策略文件失败: {e}")
                self.statusBar.showMessage(f"创建策略失败: {str(e)}")

        cancel_button.clicked.connect(on_cancel)
        create_button.clicked.connect(on_create)

        # 显示对话框
        dialog.exec_()

    def start_strategy(self):
        """
        启动策略
        """
        # 获取选中的行
        selected_rows = self.strategy_list.selectionModel().selectedRows()
        if not selected_rows:
            self.statusBar.showMessage("请选择要启动的策略")
            return

        # 启动选中的策略
        for row in selected_rows:
            strategy_name = self.strategy_list.item(row.row(), 0).text()

            # 检查策略是否已经在运行
            if (
                strategy_name in self.strategy_threads
                and self.strategy_threads[strategy_name].is_alive()
            ):
                self.statusBar.showMessage(f"策略 {strategy_name} 已经在运行")
                continue

            # 查找策略
            strategy_info = None
            for strategy in self.strategies:
                if strategy["name"] == strategy_name:
                    strategy_info = strategy
                    break

            if not strategy_info:
                self.statusBar.showMessage(f"策略 {strategy_name} 不存在")
                continue

            try:
                # 动态加载策略模块
                import os
                import importlib.util

                strategies_dir = "d:/Projects/okx_trading_bot/strategies"
                strategy_file = os.path.join(strategies_dir, f"{strategy_name}.py")

                if not os.path.exists(strategy_file):
                    self.statusBar.showMessage(f"策略文件不存在: {strategy_file}")
                    continue

                # 加载模块
                spec = importlib.util.spec_from_file_location(
                    strategy_name, strategy_file
                )
                if spec and spec.loader:
                    strategy_module = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(strategy_module)

                    # 查找策略类
                    strategy_class = None
                    for name, obj in strategy_module.__dict__.items():
                        if isinstance(obj, type) and hasattr(obj, "execute"):
                            strategy_class = obj
                            break

                    if not strategy_class:
                        self.statusBar.showMessage(
                            f"策略文件中未找到策略类: {strategy_name}"
                        )
                        continue

                    # 创建策略实例
                    strategy_instance = strategy_class()
                    self.strategy_instances[strategy_name] = strategy_instance

                    # 启动策略
                    strategy_instance.start()

                    # 更新策略状态
                    strategy_info["status"] = "运行中"
                    # 添加状态颜色指示
                    status_item = QTableWidgetItem("运行中")
                    status_item.setForeground(QColor("green"))
                    self.strategy_list.setItem(row.row(), 2, status_item)

                    # 启动监控线程
                    import threading

                    def run_strategy():
                        import time

                        while strategy_info["status"] == "运行中":
                            try:
                                # 模拟市场数据
                                market_data = {
                                    "inst_id": "BTC-USDT-SWAP",
                                    "price": 40000.0,
                                    "timestamp": time.time(),
                                }
                                # 执行策略
                                strategy_instance.execute(market_data)
                                time.sleep(1)  # 每秒执行一次
                            except Exception as e:
                                logger.error(f"策略执行错误: {e}")
                                time.sleep(1)

                    # 启动线程
                    thread = threading.Thread(target=run_strategy, daemon=True)
                    thread.start()
                    self.strategy_threads[strategy_name] = thread

                    # 更新策略列表
                    self.update_strategy_list()
                    # 更新性能图表
                    self.update_performance_chart()

                    self.statusBar.showMessage(f"策略 {strategy_name} 启动成功")
            except Exception as e:
                logger.error(f"启动策略错误: {e}")
                self.statusBar.showMessage(f"启动策略失败: {str(e)}")

    def stop_strategy(self):
        """
        停止策略
        """
        # 获取选中的行
        selected_rows = self.strategy_list.selectionModel().selectedRows()
        if not selected_rows:
            self.statusBar.showMessage("请选择要停止的策略")
            return

        # 停止选中的策略
        for row in selected_rows:
            strategy_name = self.strategy_list.item(row.row(), 0).text()

            # 查找策略
            strategy_info = None
            for strategy in self.strategies:
                if strategy["name"] == strategy_name:
                    strategy_info = strategy
                    break

            if not strategy_info:
                self.statusBar.showMessage(f"策略 {strategy_name} 不存在")
                continue

            # 更新策略状态
            strategy_info["status"] = "已停止"
            # 添加状态颜色指示
            status_item = QTableWidgetItem("已停止")
            status_item.setForeground(QColor("gray"))
            self.strategy_list.setItem(row.row(), 2, status_item)

            # 停止策略实例
            if strategy_name in self.strategy_instances:
                try:
                    strategy_instance = self.strategy_instances[strategy_name]
                    if hasattr(strategy_instance, "stop"):
                        strategy_instance.stop()
                except Exception as e:
                    logger.error(f"停止策略实例错误: {e}")

            # 清理线程
            if strategy_name in self.strategy_threads:
                del self.strategy_threads[strategy_name]

            # 更新策略列表
            self.update_strategy_list()
            # 更新性能图表
            self.update_performance_chart()

            self.statusBar.showMessage(f"策略 {strategy_name} 已停止")

    def edit_strategy(self):
        """
        编辑策略
        """
        # 获取选中的行
        selected_rows = self.strategy_list.selectionModel().selectedRows()
        if not selected_rows:
            self.statusBar.showMessage("请选择要编辑的策略")
            return

        # 只编辑第一个选中的策略
        row = selected_rows[0]
        strategy_name = self.strategy_list.item(row.row(), 0).text()

        # 查找策略
        strategy_info = None
        for strategy in self.strategies:
            if strategy["name"] == strategy_name:
                strategy_info = strategy
                break

        if not strategy_info:
            self.statusBar.showMessage(f"策略 {strategy_name} 不存在")
            return

        # 打开代码编辑对话框
        from PyQt5.QtWidgets import (
            QDialog,
            QVBoxLayout,
            QTextEdit,
            QPushButton,
            QLabel,
            QHBoxLayout,
        )

        dialog = QDialog(self)
        dialog.setWindowTitle(f"编辑策略: {strategy_name}")
        dialog.setGeometry(200, 200, 800, 600)

        layout = QVBoxLayout(dialog)

        # 提示信息
        info_label = QLabel("编辑策略代码:")
        layout.addWidget(info_label)

        # 代码输入区域
        code_editor = QTextEdit()
        
        # 加载现有代码
        import os
        strategies_dir = "d:/Projects/okx_trading_bot/strategies"
        strategy_file = os.path.join(strategies_dir, f"{strategy_name}.py")
        
        try:
            with open(strategy_file, "r", encoding="utf-8") as f:
                code = f.read()
            code_editor.setPlainText(code)
        except Exception as e:
            logger.error(f"读取策略文件错误: {e}")
            code_editor.setPlaceholderText("无法加载策略代码")

        layout.addWidget(code_editor)

        # 按钮布局
        button_layout = QHBoxLayout()
        cancel_button = QPushButton("取消")
        save_button = QPushButton("保存")

        button_layout.addStretch(1)
        button_layout.addWidget(cancel_button)
        button_layout.addWidget(save_button)

        layout.addLayout(button_layout)

        # 连接按钮信号
        def on_cancel():
            dialog.reject()

        def on_save():
            code = code_editor.toPlainText().strip()
            if not code:
                self.statusBar.showMessage("策略代码不能为空")
                return

            # 保存策略代码到文件
            try:
                with open(strategy_file, "w", encoding="utf-8") as f:
                    f.write(code)
                logger.info(f"策略文件已更新: {strategy_file}")

                self.statusBar.showMessage(f"策略 {strategy_name} 编辑成功")
                dialog.accept()
            except Exception as e:
                logger.error(f"保存策略文件失败: {e}")
                self.statusBar.showMessage(f"编辑策略失败: {str(e)}")

        cancel_button.clicked.connect(on_cancel)
        save_button.clicked.connect(on_save)

        # 显示对话框
        dialog.exec_()

    def delete_strategy(self):
        """
        删除策略
        """
        # 获取选中的行
        selected_rows = self.strategy_list.selectionModel().selectedRows()
        if not selected_rows:
            self.statusBar.showMessage("请选择要删除的策略")
            return

        # 确认删除
        from PyQt5.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self,
            "确认删除",
            "确定要删除选中的策略吗？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        # 删除选中的策略
        for row in selected_rows:
            strategy_name = self.strategy_list.item(row.row(), 0).text()

            # 查找策略
            strategy_info = None
            for i, strategy in enumerate(self.strategies):
                if strategy["name"] == strategy_name:
                    strategy_info = strategy
                    del self.strategies[i]
                    break

            if not strategy_info:
                continue

            # 停止策略（如果正在运行）
            if strategy_info["status"] == "运行中":
                # 停止策略实例
                if strategy_name in self.strategy_instances:
                    try:
                        strategy_instance = self.strategy_instances[strategy_name]
                        if hasattr(strategy_instance, "stop"):
                            strategy_instance.stop()
                    except Exception as e:
                        logger.error(f"停止策略实例错误: {e}")

                # 清理线程
                if strategy_name in self.strategy_threads:
                    del self.strategy_threads[strategy_name]

            # 删除策略文件
            import os
            strategies_dir = "d:/Projects/okx_trading_bot/strategies"
            strategy_file = os.path.join(strategies_dir, f"{strategy_name}.py")
            
            try:
                if os.path.exists(strategy_file):
                    os.remove(strategy_file)
                    logger.info(f"策略文件已删除: {strategy_file}")
            except Exception as e:
                logger.error(f"删除策略文件错误: {e}")

            # 清理策略实例
            if strategy_name in self.strategy_instances:
                del self.strategy_instances[strategy_name]

        # 更新策略列表
        self.update_strategy_list()
        # 更新性能图表
        self.update_performance_chart()

        self.statusBar.showMessage("策略删除成功")

    def update_strategy_list(self):
        """
        更新策略列表
        """
        # 清空表格
        self.strategy_list.setRowCount(0)

        # 添加策略到表格
        for strategy in self.strategies:
            row_position = self.strategy_list.rowCount()
            self.strategy_list.insertRow(row_position)
            
            # 策略名称
            self.strategy_list.setItem(row_position, 0, QTableWidgetItem(strategy["name"]))
            
            # 策略类型
            self.strategy_list.setItem(row_position, 1, QTableWidgetItem(strategy["type"]))
            
            # 策略状态（带颜色）
            status_item = QTableWidgetItem(strategy["status"])
            if strategy["status"] == "运行中":
                status_item.setForeground(QColor("green"))
            else:
                status_item.setForeground(QColor("gray"))
            self.strategy_list.setItem(row_position, 2, status_item)
            
            # 策略性能（模拟数据）
            performance_item = QTableWidgetItem("+12.5%")
            performance_item.setForeground(QColor("green"))
            self.strategy_list.setItem(row_position, 3, performance_item)
            
            # 操作按钮
            button_widget = QWidget()
            button_layout = QHBoxLayout(button_widget)
            button_layout.setContentsMargins(0, 0, 0, 0)
            button_layout.setSpacing(5)
            
            # 启动按钮
            start_btn = QPushButton("启动")
            start_btn.setFixedSize(60, 24)
            start_btn.setStyleSheet("font-size: 12px; padding: 2px;")
            start_btn.clicked.connect(lambda checked, name=strategy["name"]: self._start_strategy_by_name(name))
            
            # 停止按钮
            stop_btn = QPushButton("停止")
            stop_btn.setFixedSize(60, 24)
            stop_btn.setStyleSheet("font-size: 12px; padding: 2px;")
            stop_btn.clicked.connect(lambda checked, name=strategy["name"]: self._stop_strategy_by_name(name))
            
            # 编辑按钮
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedSize(60, 24)
            edit_btn.setStyleSheet("font-size: 12px; padding: 2px;")
            edit_btn.clicked.connect(lambda checked, name=strategy["name"]: self._edit_strategy_by_name(name))
            
            # 删除按钮
            delete_btn = QPushButton("删除")
            delete_btn.setFixedSize(60, 24)
            delete_btn.setStyleSheet("font-size: 12px; padding: 2px; background-color: #f44336; color: white;")
            delete_btn.clicked.connect(lambda checked, name=strategy["name"]: self._delete_strategy_by_name(name))
            
            button_layout.addWidget(start_btn)
            button_layout.addWidget(stop_btn)
            button_layout.addWidget(edit_btn)
            button_layout.addWidget(delete_btn)
            button_layout.addStretch(1)
            
            self.strategy_list.setCellWidget(row_position, 4, button_widget)

    def _start_strategy_by_name(self, strategy_name):
        """
        根据名称启动策略
        """
        # 查找策略在表格中的行
        for row in range(self.strategy_list.rowCount()):
            if self.strategy_list.item(row, 0).text() == strategy_name:
                # 模拟选择该行
                self.strategy_list.selectRow(row)
                # 调用启动策略方法
                self.start_strategy()
                break

    def _stop_strategy_by_name(self, strategy_name):
        """
        根据名称停止策略
        """
        # 查找策略在表格中的行
        for row in range(self.strategy_list.rowCount()):
            if self.strategy_list.item(row, 0).text() == strategy_name:
                # 模拟选择该行
                self.strategy_list.selectRow(row)
                # 调用停止策略方法
                self.stop_strategy()
                break

    def _edit_strategy_by_name(self, strategy_name):
        """
        根据名称编辑策略
        """
        # 查找策略在表格中的行
        for row in range(self.strategy_list.rowCount()):
            if self.strategy_list.item(row, 0).text() == strategy_name:
                # 模拟选择该行
                self.strategy_list.selectRow(row)
                # 调用编辑策略方法
                self.edit_strategy()
                break

    def _delete_strategy_by_name(self, strategy_name):
        """
        根据名称删除策略
        """
        # 查找策略在表格中的行
        for row in range(self.strategy_list.rowCount()):
            if self.strategy_list.item(row, 0).text() == strategy_name:
                # 模拟选择该行
                self.strategy_list.selectRow(row)
                # 调用删除策略方法
                self.delete_strategy()
                break

    def update_performance_chart(self):
        """
        更新策略性能图表
        """
        try:
            # 清空图表
            self.performance_figure.clear()
            
            # 模拟性能数据
            if self.strategies:
                # 创建示例数据
                import numpy as np
                import matplotlib.pyplot as plt
                from datetime import datetime, timedelta
                
                # 生成最近24小时的时间数据
                now = datetime.now()
                times = [now - timedelta(hours=i) for i in range(24)][::-1]
                
                # 为每个策略生成模拟性能数据
                ax = self.performance_figure.add_subplot(111)
                
                colors = ['blue', 'green', 'red', 'purple', 'orange']
                
                for i, strategy in enumerate(self.strategies[:5]):  # 最多显示5个策略
                    # 生成模拟的累计收益数据
                    base_profit = np.random.uniform(0, 10)
                    noise = np.random.normal(0, 0.5, len(times))
                    profits = base_profit + np.cumsum(noise)
                    
                    # 绘制曲线
                    ax.plot(times, profits, label=strategy['name'], color=colors[i % len(colors)], linewidth=2)
                
                # 设置图表标题和标签
                ax.set_title("策略性能对比")
                ax.set_xlabel("时间")
                ax.set_ylabel("累计收益 (%)")
                
                # 设置x轴格式
                import matplotlib.dates as mdates
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
                self.performance_figure.autofmt_xdate()
                
                # 添加网格和图例
                ax.grid(True, linestyle='--', alpha=0.7)
                ax.legend()
            else:
                # 没有策略时显示提示
                ax = self.performance_figure.add_subplot(111)
                ax.set_title("策略性能对比")
                ax.set_xlabel("时间")
                ax.set_ylabel("累计收益 (%)")
                ax.text(0.5, 0.5, "暂无策略数据", ha='center', va='center', transform=ax.transAxes)
            
            # 重新绘制图表
            self.performance_canvas.draw()
        except Exception as e:
            logger.error(f"更新性能图表错误: {e}")

    def on_market_data_update(self, data):
        """
        更新市场数据
        """
        channel = data.get("channel")
        inst_id = data.get("inst_id")
        market_data = data.get("data", [])

        if channel == "tickers" and market_data:
            ticker_data = market_data[0]
            last_price = ticker_data.get("last", "0")
            change24h = ticker_data.get("change24h", "0")
            change24h_percent = ticker_data.get("change24hPercent", "0")
            vol24h = ticker_data.get("vol24h", "0")

            # 更新市场数据存储
            self.market_data[inst_id] = {
                "last": last_price,
                "change24h": change24h,
                "change24hPercent": change24h_percent,
                "vol24h": vol24h,
                "update_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }

            # 更新价格历史数据
            if inst_id not in self.price_history:
                self.price_history[inst_id] = []
            
            # 添加新的价格数据，保留最近200个数据点
            timestamp = datetime.now()
            price = float(last_price)
            self.price_history[inst_id].append((timestamp, price))
            if len(self.price_history[inst_id]) > 200:
                self.price_history[inst_id] = self.price_history[inst_id][-200:]
            
            # 更新价格图表（如果当前选中的是该产品）
            if hasattr(self, 'chart_product_combo') and self.chart_product_combo.currentText() == inst_id:
                self.update_price_chart(inst_id)
            
            # 更新市场概览
            if inst_id == "BTC-USDT-SWAP":
                self.market_price_label.setText(f"BTC/USDT: ${float(last_price):,.2f}")
                change_percent = float(change24h_percent) * 100
                if change_percent > 0:
                    self.market_change_label.setText(f"24h: +{change_percent:.2f}%")
                    self.market_change_label.setStyleSheet("font-size: 16px; color: green;")
                else:
                    self.market_change_label.setText(f"24h: {change_percent:.2f}%")
                    self.market_change_label.setStyleSheet("font-size: 16px; color: red;")
            
            # 更新市场分析指标
            self.update_market_analysis(inst_id, float(last_price), float(change24h_percent) * 100)
            
            # 更新市场数据表格
            self.update_market_table()

    def update_market_analysis(self, inst_id, price, change_percent):
        """
        更新市场分析指标
        """
        try:
            # 只更新BTC的分析数据
            if inst_id == 'BTC-USDT-SWAP' and hasattr(self, 'trend_label'):
                # 更新趋势分析
                if abs(change_percent) > 2:
                    if change_percent > 0:
                        self.trend_label.setText("当前趋势: 上涨")
                        self.trend_label.setStyleSheet("font-size: 14px; font-weight: bold; color: green;")
                    else:
                        self.trend_label.setText("当前趋势: 下跌")
                        self.trend_label.setStyleSheet("font-size: 14px; font-weight: bold; color: red;")
                else:
                    self.trend_label.setText("当前趋势: 横盘")
                    self.trend_label.setStyleSheet("font-size: 14px; font-weight: bold; color: gray;")
                
                # 更新趋势强度
                if hasattr(self, 'trend_strength_label'):
                    trend_strength = min(100, abs(change_percent) * 10)
                    self.trend_strength_label.setText(f"趋势强度: {int(trend_strength)}")
                
                # 更新支撑阻力位（模拟数据）
                support = price * 0.95
                resistance = price * 1.05
                if hasattr(self, 'support_resistance_label'):
                    self.support_resistance_label.setText(f"支撑位: ${support:,.2f} | 阻力位: ${resistance:,.2f}")
                
                # 更新市场情绪
                if hasattr(self, 'sentiment_label') and hasattr(self, 'sentiment_value_label'):
                    if change_percent > 3:
                        self.sentiment_label.setText("市场情绪: 极度乐观")
                        self.sentiment_label.setStyleSheet("font-size: 14px; font-weight: bold; color: green;")
                        self.sentiment_value_label.setText("情绪指数: 90/100")
                    elif change_percent > 1:
                        self.sentiment_label.setText("市场情绪: 乐观")
                        self.sentiment_label.setStyleSheet("font-size: 14px; font-weight: bold; color: green;")
                        self.sentiment_value_label.setText("情绪指数: 75/100")
                    elif change_percent < -3:
                        self.sentiment_label.setText("市场情绪: 极度悲观")
                        self.sentiment_label.setStyleSheet("font-size: 14px; font-weight: bold; color: red;")
                        self.sentiment_value_label.setText("情绪指数: 10/100")
                    elif change_percent < -1:
                        self.sentiment_label.setText("市场情绪: 悲观")
                        self.sentiment_label.setStyleSheet("font-size: 14px; font-weight: bold; color: red;")
                        self.sentiment_value_label.setText("情绪指数: 25/100")
                    else:
                        self.sentiment_label.setText("市场情绪: 中性")
                        self.sentiment_label.setStyleSheet("font-size: 14px; font-weight: bold; color: gray;")
                        self.sentiment_value_label.setText("情绪指数: 50/100")
                
                # 更新恐慌与贪婪指数
                if hasattr(self, 'fear_greed_label'):
                    if change_percent > 3:
                        fear_greed = 90
                    elif change_percent > 1:
                        fear_greed = 75
                    elif change_percent < -3:
                        fear_greed = 10
                    elif change_percent < -1:
                        fear_greed = 25
                    else:
                        fear_greed = 50
                    self.fear_greed_label.setText(f"恐慌与贪婪指数: {fear_greed}")
                
                # 更新波动率分析
                if hasattr(self, 'volatility_label') and hasattr(self, 'volatility_value_label'):
                    volatility = abs(change_percent)
                    if volatility > 5:
                        self.volatility_label.setText("波动率: 高")
                        self.volatility_label.setStyleSheet("font-size: 14px; font-weight: bold; color: red;")
                        self.volatility_value_label.setText(f"24h波动率: {volatility:.2f}%")
                    elif volatility > 2:
                        self.volatility_label.setText("波动率: 中")
                        self.volatility_label.setStyleSheet("font-size: 14px; font-weight: bold; color: orange;")
                        self.volatility_value_label.setText(f"24h波动率: {volatility:.2f}%")
                    else:
                        self.volatility_label.setText("波动率: 低")
                        self.volatility_label.setStyleSheet("font-size: 14px; font-weight: bold; color: green;")
                        self.volatility_value_label.setText(f"24h波动率: {volatility:.2f}%")
                
                # 更新ATR值（模拟数据）
                if hasattr(self, 'atr_label'):
                    atr = price * 0.01 * (volatility / 2.5)
                    self.atr_label.setText(f"ATR: {atr:.2f}")
                
                # 更新量价关系分析
                if hasattr(self, 'volume_trend_label') and hasattr(self, 'volume_change_label'):
                    # 模拟成交量变化
                    volume_change = change_percent * 1.5
                    if change_percent > 0 and volume_change > 0:
                        self.volume_trend_label.setText("量价趋势: 同步上涨")
                        self.volume_trend_label.setStyleSheet("color: green;")
                    elif change_percent < 0 and volume_change < 0:
                        self.volume_trend_label.setText("量价趋势: 同步下跌")
                        self.volume_trend_label.setStyleSheet("color: red;")
                    else:
                        self.volume_trend_label.setText("量价趋势: 背离")
                        self.volume_trend_label.setStyleSheet("color: orange;")
                    self.volume_change_label.setText(f"成交量变化: {volume_change:+.2f}%")
                
                # 更新市场广度分析
                if hasattr(self, 'advance_decline_label') and hasattr(self, 'market_breadth_label'):
                    # 模拟市场广度数据
                    if change_percent > 1:
                        advance_decline = 1.5
                        market_breadth = 65
                    elif change_percent < -1:
                        advance_decline = 0.5
                        market_breadth = 35
                    else:
                        advance_decline = 1.0
                        market_breadth = 50
                    self.advance_decline_label.setText(f"涨跌比: {advance_decline:.1f}")
                    self.market_breadth_label.setText(f"市场广度: {market_breadth}%")
                
                # 更新技术形态识别
                if hasattr(self, 'pattern_label') and hasattr(self, 'pattern_strength_label'):
                    # 模拟技术形态识别
                    if abs(change_percent) > 3:
                        if change_percent > 0:
                            self.pattern_label.setText("当前形态: 突破")
                        else:
                            self.pattern_label.setText("当前形态: 跌破")
                        pattern_strength = 80
                    elif abs(change_percent) > 1:
                        if change_percent > 0:
                            self.pattern_label.setText("当前形态: 上升趋势")
                        else:
                            self.pattern_label.setText("当前形态: 下降趋势")
                        pattern_strength = 60
                    else:
                        self.pattern_label.setText("当前形态: 整理")
                        pattern_strength = 40
                    self.pattern_strength_label.setText(f"形态强度: {pattern_strength}")
        except Exception as e:
            logger.error(f"更新市场分析错误: {e}")

    def on_order_data_update(self, data):
        """
        更新订单数据
        """
        try:
            if 'data' in data:
                for item in data['data']:
                    order_id = item.get('ordId', '')
                    inst_id = item.get('instId', '')
                    side = item.get('side', '')
                    ord_type = item.get('ordType', '')
                    price = item.get('price', '0')
                    size = item.get('size', '0')
                    status = item.get('state', '')
                    update_time = item.get('uTime', '')
                    
                    # 更新订单数据存储
                    self.order_data[order_id] = {
                        'order_id': order_id,
                        'inst_id': inst_id,
                        'side': side,
                        'type': ord_type,
                        'price': price,
                        'size': size,
                        'status': status,
                        'update_time': update_time,
                    }
                    
                    # 更新订单表格
                    self.update_order_table()
        except Exception as e:
            logger.error(f"处理订单数据更新错误: {e}")

    def on_account_data_update(self, data):
        """
        更新账户数据
        """
        try:
            if 'data' in data:
                for item in data['data']:
                    if 'balance' in item:
                        for balance in item['balance']:
                            ccy = balance.get('ccy', '')
                            avail_balance = balance.get('availBal', '0')
                            frozen_balance = balance.get('frozenBal', '0')
                            total_balance = balance.get('totalBal', '0')
                            
                            # 更新账户数据存储
                            self.account_data[ccy] = {
                                'ccy': ccy,
                                'avail_balance': avail_balance,
                                'frozen_balance': frozen_balance,
                                'total_balance': total_balance,
                            }
                            
                            # 更新账户表格
                            self.update_account_table()
        except Exception as e:
            logger.error(f"处理账户数据更新错误: {e}")

    def on_connection_status_update(self, connected, message):
        """
        更新连接状态
        """
        if connected:
            self.connection_status.setText("已连接")
            self.connection_status.setStyleSheet("color: green; font-weight: bold; font-size: 14px;")
        else:
            self.connection_status.setText("未连接")
            self.connection_status.setStyleSheet("color: red; font-weight: bold; font-size: 14px;")

    def on_market_event(self, event):
        """
        处理市场事件
        """
        self.update_market_data.emit(event.data)

    def on_order_event(self, event):
        """
        处理订单事件
        """
        self.update_order_data.emit(event.data)

    def on_account_event(self, event):
        """
        处理账户事件
        """
        self.update_account_data.emit(event.data)

    def on_ws_connected(self, event):
        """
        处理WebSocket连接事件
        """
        self.update_connection_status.emit(True, "已连接")

    def on_ws_disconnected(self, event):
        """
        处理WebSocket断开连接事件
        """
        self.update_connection_status.emit(False, "已断开")

    def export_market_data_csv(self):
        """
        导出市场数据为CSV格式
        """
        try:
            # 获取保存文件路径
            from PyQt5.QtWidgets import QFileDialog
            file_path, _ =FileDialog.getSaveFileName(
                self,
                "保存CSV文件",
                "",
                "市场数据.csv",
                "CSV Files (*.csv);;*.csv"
            )
            
            if not file_path:
                return
            
            # 准备数据
            import csv
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                # 写入表头
                writer = csv.writer(csvfile)
                writer.writerow(['交易对', '最新价格', '涨跌幅', '成交量', '最高价', '最低价', '更新时间'])
                
                # 写入数据
                for inst_id, data in self.market_data.items():
                    writer.writerow([
                        inst_id,
                        data.get('last', '0'),
                        data.get('change24hPercent', '0'),
                        data.get('vol24h', '0'),
                        data.get('last', '0'),  # 使用last作为最高价
                        data.get('last', '0'),  # 使用last作为最低价
                        data.get('update_time', '')
                    ])
            
            self.statusBar.showMessage(f"市场数据已导出到: {file_path}")
            logger.info(f"市场数据已导出到: {file_path}")
            
        except Exception as e:
            logger.error(f"导出市场数据错误: {e}")
            QMessageBox.critical(self, "错误", f"导出市场数据失败: {str(e)}")

    def export_market_data_excel(self):
        """
        导出市场数据为Excel格式
        """
        try:
            # 检查是否安装了openpyxl
            try:
                import openpyxl
                from openpyxl import Workbook
            except ImportError:
                QMessageBox.warning(self, "警告", "需要安装openpyxl库才能导出Excel文件。\n请运行: pip install openpyxl")
                return
            
            # 获取保存文件路径
            from PyQt5.QtWidgets import QFileDialog
            file_path, _ =FileDialog.getSaveFileName(
                self,
                "保存Excel文件",
                "",
                "市场数据.xlsx",
                "Excel Files (*.xlsx);;*.xlsx"
            )
            
            if not file_path:
                return
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            
            
            # 写入表头
            headers = ['交易对', '最新价格', '涨跌幅', '成交量', '最高价', '最低价', '更新时间']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            row = 2
            for inst_id, data in self.market_data.items():
                ws.cell(row=row, column=1, value=inst_id)
                ws.cell(row=row, column=2, value=data.get('last', '0'))
                ws.cell(row=row, column=3, value=data.get('change24hPercent', '0'))
                ws.cell(row=row, column=4, value=data.get('vol24h', '0'))
                ws.cell(row=row, column=5, value=data.get('last', '0'))  # 使用last作为最高价
                ws.cell(row=row, column=6, value=data.get('last', '0'))  # 使用last作为最低价
                ws.cell(row=row, column=7, value=data.get('update_time', ''))
                row += 1
            
            # 保存文件
            wb.save(file_path)
            
            self.statusBar.showMessage(f"市场数据已导出到: {file_path}")
            logger.info(f"市场数据已导出到: {file_path}")
            
        except Exception as e:
            logger.error(f"导出市场数据错误: {e}")
            QMessageBox.critical(self, "错误", f"导出市场数据失败: {str(e)}")

    def init_help_menu(self):
        """
        初始化帮助菜单
        """
        from PyQt5.QtWidgets import QMenuBar, QAction
        
        menubar = self.menuBar()
        help_menu = menubar.addMenu('帮助')
        
        about_action = QAction('关于', self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
        
        help_action = QAction('使用帮助', self)
        help_action.triggered.connect(self.show_help)
        help_menu.addAction(help_action)
        
        # 添加主题切换菜单
        view_menu = menubar.addMenu('视图')
        
        theme_action = QAction('切换主题', self)
        theme_action.triggered.connect(self.toggle_theme)
        view_menu.addAction(theme_action)
        
        # 启动仪表盘更新定时器
        self.start_dashboard_timer()

    def start_dashboard_timer(self):
        """
        启动仪表盘更新定时器
        """
        # 创建定时器，每5秒更新一次仪表盘数据
        self.dashboard_timer = QTimer(self)
        self.dashboard_timer.timeout.connect(self.update_dashboard)
        self.dashboard_timer.start(5000)  # 5秒更新一次
        logger.info("仪表盘更新定时器已启动")

    def update_dashboard(self):
        """
        更新仪表盘数据
        """
        try:
            # 更新市场概览
            self.update_market_overview()
            
            # 更新账户概览
            self.update_account_overview()
            
            # 更新策略概览
            self.update_strategy_overview()
            
            # 更新交易概览
            self.update_trade_overview()
            
            # 更新风险指标
            self.update_risk_overview()
            
            # 更新持仓概览
            self.update_position_overview()
        except Exception as e:
            logger.error(f"更新仪表盘错误: {e}")

    def update_market_overview(self):
        """
        更新市场概览
        """
        try:
            # 获取BTC-USDT-SWAP的市场数据
            if "BTC-USDT-SWAP" in self.market_data:
                data = self.market_data["BTC-USDT-SWAP"]
                last_price = float(data.get("last", 0))
                change_percent = float(data.get("change24hPercent", 0)) * 100
                
                # 更新价格标签
                self.market_price_label.setText(f"BTC/USDT: ${last_price:,.2f}")
                
                # 更新涨跌幅标签
                if change_percent > 0:
                    self.market_change_label.setText(f"24h: +{change_percent:.2f}%")
                    self.market_change_label.setStyleSheet("font-size: 16px; color: green;")
                else:
                    self.market_change_label.setText(f"24h: {change_percent:.2f}%")
                    self.market_change_label.setStyleSheet("font-size: 16px; color: red;")
        except Exception as e:
            logger.error(f"更新市场概览错误: {e}")

    def update_account_overview(self):
        """
        更新账户概览
        """
        try:
            # 计算总余额
            total_balance = 0.0
            if isinstance(self.account_data, dict):
                for ccy, data in self.account_data.items():
                    total_balance += float(data.get('total_balance', 0))
            
            # 更新余额标签
            self.account_balance_label.setText(f"总余额: ${total_balance:,.2f}")
            
            # 计算今日盈亏（模拟数据）
            # 在实际应用中，应该从交易记录中计算
            today_pnl = 0
            if isinstance(self.order_data, dict):
                today_pnl = sum(1 for order in self.order_data.values() 
                              if order.get('update_time', '').startswith(datetime.now().strftime('%Y-%m-%d')))
            elif isinstance(self.order_data, list):
                today_pnl = sum(1 for order in self.order_data 
                              if order.get('update_time', '').startswith(datetime.now().strftime('%Y-%m-%d')))
            
            # 更新盈亏标签
            if today_pnl > 0:
                self.account_pnl_label.setText(f"今日盈亏: +${today_pnl * 10:.2f}")
                self.account_pnl_label.setStyleSheet("font-size: 16px; color: green;")
            else:
                self.account_pnl_label.setText(f"今日盈亏: ${today_pnl * 10:.2f}")
                self.account_pnl_label.setStyleSheet("font-size: 16px; color: red;")
        except Exception as e:
            logger.error(f"更新账户概览错误: {e}")

    def update_strategy_overview(self):
        """
        更新策略概览
        """
        try:
            # 统计运行中的策略数量
            running_strategies = sum(1 for s in self.strategies if s.get('status') == '运行中')
            
            # 更新策略状态标签
            self.strategy_status_label.setText(f"运行中: {running_strategies}个策略")
            
            # 计算总利润（模拟数据）
            # 在实际应用中，应该从策略执行结果中计算
            total_profit = running_strategies * 100
            
            # 更新利润标签
            if total_profit > 0:
                self.strategy_profit_label.setText(f"总利润: +${total_profit:.2f}")
                self.strategy_profit_label.setStyleSheet("font-size: 16px; color: green;")
            else:
                self.strategy_profit_label.setText(f"总利润: ${total_profit:.2f}")
                self.strategy_profit_label.setStyleSheet("font-size: 16px; color: red;")
        except Exception as e:
            logger.error(f"更新策略概览错误: {e}")

    def update_trade_overview(self):
        """
        更新交易概览
        """
        try:
            # 统计今日交易数量
            today_trades = 0
            if isinstance(self.order_data, dict):
                today_trades = sum(1 for order in self.order_data.values() 
                                 if order.get('update_time', '').startswith(datetime.now().strftime('%Y-%m-%d')))
            elif isinstance(self.order_data, list):
                today_trades = sum(1 for order in self.order_data 
                                 if order.get('update_time', '').startswith(datetime.now().strftime('%Y-%m-%d')))
            
            # 更新交易数量标签
            self.trade_count_label.setText(f"今日交易: {today_trades}笔")
            
            # 计算胜率（模拟数据）
            # 在实际应用中，应该从交易记录中计算
            win_rate = 75 if today_trades > 0 else 0
            
            # 更新胜率标签
            self.trade_win_rate_label.setText(f"胜率: {win_rate}%")
            if win_rate > 50:
                self.trade_win_rate_label.setStyleSheet("font-size: 16px; color: green;")
            else:
                self.trade_win_rate_label.setStyleSheet("font-size: 16px; color: red;")
        except Exception as e:
            logger.error(f"更新交易概览错误: {e}")
    
    def update_risk_overview(self):
        """
        更新风险指标
        """
        try:
            # 计算风险等级（模拟数据）
            # 在实际应用中，应该根据策略表现计算
            risk_level = "中等"
            max_drawdown = "-3.2%"
            
            # 更新风险等级标签
            self.risk_level_label.setText(f"风险等级: {risk_level}")
            
            # 更新最大回撤标签
            self.max_drawdown_label.setText(f"最大回撤: {max_drawdown}")
            self.max_drawdown_label.setStyleSheet("font-size: 16px; color: red;")
        except Exception as e:
            logger.error(f"更新风险指标错误: {e}")
    
    def update_position_overview(self):
        """
        更新持仓概览
        """
        try:
            # 统计持仓数量（模拟数据）
            # 在实际应用中，应该从账户数据中计算
            position_count = 3
            position_value = 5000
            
            # 更新持仓数量标签
            self.position_count_label.setText(f"持仓数量: {position_count}个")
            
            # 更新持仓价值标签
            self.position_value_label.setText(f"持仓价值: ${position_value:,.2f}")
            self.position_value_label.setStyleSheet("font-size: 16px; color: green;")
        except Exception as e:
            logger.error(f"更新持仓概览错误: {e}")

    def generate_strategy_report(self):
        """
        生成策略性能分析报告
        """
        try:
            # 获取选中的策略
            selected_rows = self.strategy_list.selectionModel().selectedRows()
            if not selected_rows:
                self.statusBar.showMessage("请选择要生成报告的策略")
                return
            
            # 只处理第一个选中的策略
            row = selected_rows[0]
            strategy_name = self.strategy_list.item(row.row(), 0).text()
            
            # 查找策略
            strategy_info = None
            for strategy in self.strategies:
                if strategy["name"] == strategy_name:
                    strategy_info = strategy
                    break
            
            if not strategy_info:
                self.statusBar.showMessage(f"策略 {strategy_name} 不存在")
                return
            
            # 生成报告
            report_content = self._generate_report_content(strategy_name, strategy_info)
            
            # 保存报告文件
            from PyQt5.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "保存策略报告",
                f"{strategy_name}_报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.md",
                "Markdown Files (*.md);;All Files (*)"
            )
            
            if not file_path:
                return
            
            # 写入报告文件
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(report_content)
            
            self.statusBar.showMessage(f"策略报告已生成: {file_path}")
            logger.info(f"策略报告已生成: {file_path}")
            
        except Exception as e:
            logger.error(f"生成策略报告错误: {e}")
            self.statusBar.showMessage(f"生成策略报告失败: {str(e)}")
    
    def _generate_report_content(self, strategy_name, strategy_info):
        """
        生成报告内容
        """
        import pandas as pd
        from datetime import datetime, timedelta
        
        # 生成报告头部
        report = f"# {strategy_name} 策略性能分析报告\n\n"
        report += f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        report += f"策略类型: {strategy_info.get('type', '未知')}\n"
        report += f"策略状态: {strategy_info.get('status', '未知')}\n\n"
        
        # 生成性能指标
        report += "## 性能指标\n\n"
        
        # 模拟性能数据
        performance_data = {
            "总收益率": "12.5%",
            "最大回撤": "-3.2%",
            "夏普比率": "2.1",
            "年化收益率": "15.3%",
            "胜率": "68%",
            "平均盈亏比": "1.8",
            "交易次数": "128",
            "平均持仓时间": "4.2小时"
        }
        
        for key, value in performance_data.items():
            report += f"- **{key}**: {value}\n"
        
        # 生成交易记录统计
        report += "\n## 交易记录统计\n\n"
        
        # 模拟交易数据
        trade_data = []
        for i in range(10):
            date = (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d')
            trade_data.append({
                "日期": date,
                "交易次数": str(10 + i),
                "盈利次数": str(7 + i % 3),
                "亏损次数": str(3 - i % 3),
                "当日收益": f"{((i % 5) - 2) * 0.5}%"
            })
        
        # 转换为表格
        df = pd.DataFrame(trade_data)
        report += df.to_markdown(index=False) + "\n\n"
        
        # 生成风险分析
        report += "## 风险分析\n\n"
        report += "- **风险等级**: 中等\n"
        report += "- **最大连续亏损**: 3次\n"
        report += "- **最大连续盈利**: 5次\n"
        report += "- **波动性**: 低\n\n"
        
        # 生成策略建议
        report += "## 策略建议\n\n"
        report += "1. **参数优化**: 考虑调整止损参数，减少最大回撤\n"
        report += "2. **资金管理**: 建议将单次交易资金控制在总资金的5%以内\n"
        report += "3. **市场适应**: 在高波动市场中可适当降低交易频率\n"
        report += "4. **监控建议**: 定期检查策略表现，每两周生成一次分析报告\n\n"
        
        # 生成总结
        report += "## 总结\n\n"
        report += f"{strategy_name} 策略整体表现良好，在过去的测试期内实现了 {performance_data['总收益率']} 的总收益，"
        report += f"最大回撤控制在 {performance_data['最大回撤']} 以内，夏普比率 {performance_data['夏普比率']} 表明策略风险调整后收益表现优秀。\n"
        report += "建议继续运行该策略，并根据市场变化适时调整参数以保持策略的适应性。\n"
        
        return report

    def toggle_theme(self):
        """
        切换主题
        """
        # 获取当前主题
        current_style = self.styleSheet()
        
        # 切换主题
        if "background-color: #f0f2f5" in current_style:
            # 切换到深色主题
            self.setStyleSheet("""
                QMainWindow {
                    background-color: #1e1e1e;
                }
                QGroupBox {
                    border: 1px solid #3e3e3e;
                    border-radius: 8px;
                    margin-top: 10px;
                    background-color: #2d2d2d;
                    color: #ffffff;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 15px;
                    padding: 0 10px 0 10px;
                    background-color: #2d2d2d;
                    font-weight: bold;
                    font-size: 14px;
                    color: #ffffff;
                }
                QPushButton {
                    background-color: #1976D2;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    min-width: 90px;
                    font-size: 14px;
                }
                QPushButton:hover {
                    background-color: #1565C0;
                }
                QPushButton:disabled {
                    background-color: #424242;
                }
                QLineEdit {
                    border: 1px solid #3e3e3e;
                    border-radius: 6px;
                    padding: 8px;
                    font-size: 14px;
                    background-color: #3e3e3e;
                    color: #ffffff;
                }
                QLineEdit:focus {
                    border: 1px solid #1976D2;
                }
                QComboBox {
                    border: 1px solid #3e3e3e;
                    border-radius: 6px;
                    padding: 8px;
                    font-size: 14px;
                    background-color: #3e3e3e;
                    color: #ffffff;
                }
                QTableWidget {
                    border: 1px solid #3e3e3e;
                    border-radius: 6px;
                    alternate-background-color: #363636;
                    background-color: #2d2d2d;
                    color: #ffffff;
                }
                QTableWidget::item:selected {
                    background-color: #1976D2;
                }
                QStatusBar {
                    background-color: #2d2d2d;
                    border-top: 1px solid #3e3e3e;
                    font-size: 12px;
                    color: #ffffff;
                }
                QLabel {
                    font-size: 14px;
                    color: #ffffff;
                }
                QTabBar::tab {
                    background-color: #2d2d2d;
                    padding: 12px 20px;
                    margin-right: 2px;
                    border-radius: 6px 6px 0 0;
                    font-size: 14px;
                    color: #ffffff;
                }
                QTabBar::tab:selected {
                    background-color: #3e3e3e;
                    border-bottom: 2px solid #1976D2;
                }
                QTabWidget::pane {
                    border: 1px solid #3e3e3e;
                    background-color: #2d2d2d;
                }
            """)
            self.statusBar.showMessage("已切换到深色主题")
        else:
            # 切换到浅色主题
            self.setStyleSheet("""
                QMainWindow {
                    background-color: #f0f2f5;
                }
                QGroupBox {
                    border: 1px solid #e0e0e0;
                    border-radius: 8px;
                    margin-top: 10px;
                    background-color: white;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 15px;
                    padding: 0 10px 0 10px;
                    background-color: white;
                    font-weight: bold;
                    font-size: 14px;
                }
                QPushButton {
                    background-color: #1976D2;
                    color: white;
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    min-width: 90px;
                    font-size: 14px;
                }
                QPushButton:hover {
                    background-color: #1565C0;
                }
                QPushButton:disabled {
                    background-color: #BDBDBD;
                }
                QLineEdit {
                    border: 1px solid #e0e0e0;
                    border-radius: 6px;
                    padding: 8px;
                    font-size: 14px;
                }
                QLineEdit:focus {
                    border: 1px solid #1976D2;
                }
                QComboBox {
                    border: 1px solid #e0e0e0;
                    border-radius: 6px;
                    padding: 8px;
                    font-size: 14px;
                }
                QTableWidget {
                    border: 1px solid #e0e0e0;
                    border-radius: 6px;
                    alternate-background-color: #f9f9f9;
                }
                QTableWidget::item:selected {
                    background-color: #E3F2FD;
                }
                QStatusBar {
                    background-color: white;
                    border-top: 1px solid #e0e0e0;
                    font-size: 12px;
                }
                QLabel {
                    font-size: 14px;
                }
                QTabBar::tab {
                    background-color: #f5f5f5;
                    padding: 12px 20px;
                    margin-right: 2px;
                    border-radius: 6px 6px 0 0;
                    font-size: 14px;
                }
                QTabBar::tab:selected {
                    background-color: white;
                    border-bottom: 2px solid #1976D2;
                }
            """)
            self.statusBar.showMessage("已切换到浅色主题")

    def show_about(self):
        """
        显示关于对话框
        """
        from PyQt5.QtWidgets import QMessageBox
        QMessageBox.about(self, "关于", "OKX 交易机器人\n版本: 1.0.0\n作者: OKX Trading Bot Team")

    def show_help(self):
        """
        显示帮助对话框
        """
        from PyQt5.QtWidgets import QMessageBox
        QMessageBox.information(self, "使用帮助", "1. 输入API密钥和密码\n2. 选择模拟盘或实盘\n3. 点击连接按钮\n4. 在市场数据标签页查看行情\n5. 在交易与订单标签页进行交易\n6. 在策略管理标签页管理策略\n7. 在监控与分析标签页查看监控数据")

    def update_market_table(self):
        """
        更新市场数据表格
        """
        # 清空表格
        self.market_table.setRowCount(0)
        
        # 添加市场数据到表格
        for inst_id, data in self.market_data.items():
            row_position = self.market_table.rowCount()
            self.market_table.insertRow(row_position)
            
            self.market_table.setItem(row_position, 0, QTableWidgetItem(inst_id))
            self.market_table.setItem(row_position, 1, QTableWidgetItem(f"${float(data['last']):,.2f}"))
            
            change_percent = float(data['change24hPercent']) * 100
            change_item = QTableWidgetItem(f"{change_percent:+.2f}%")
            if change_percent > 0:
                change_item.setForeground(QColor("green"))
            else:
                change_item.setForeground(QColor("red"))
            self.market_table.setItem(row_position, 2, change_item)
            
            self.market_table.setItem(row_position, 3, QTableWidgetItem(f"{float(data['vol24h']):,.2f}"))
            self.market_table.setItem(row_position, 4, QTableWidgetItem(f"${float(data['last']):,.2f}"))  # 临时使用last作为最高价
            self.market_table.setItem(row_position, 5, QTableWidgetItem(f"${float(data['last']):,.2f}"))  # 临时使用last作为最低价
            
            # 添加技术指标列
            indicator_item = QTableWidgetItem("MA: 40000")
            self.market_table.setItem(row_position, 6, indicator_item)

    def update_order_table(self):
        """
        更新订单表格
        """
        # 清空表格
        self.order_table.setRowCount(0)
        
        # 添加订单数据到表格
        for order_id, data in self.order_data.items():
            row_position = self.order_table.rowCount()
            self.order_table.insertRow(row_position)
            
            self.order_table.setItem(row_position, 0, QTableWidgetItem(data['order_id']))
            self.order_table.setItem(row_position, 1, QTableWidgetItem(data['inst_id']))
            self.order_table.setItem(row_position, 2, QTableWidgetItem(data['type']))
            self.order_table.setItem(row_position, 3, QTableWidgetItem(data['side']))
            self.order_table.setItem(row_position, 4, QTableWidgetItem(data['price']))
            self.order_table.setItem(row_position, 5, QTableWidgetItem(data['size']))
            self.order_table.setItem(row_position, 6, QTableWidgetItem(data['status']))
            self.order_table.setItem(row_position, 7, QTableWidgetItem(data['update_time']))

    def update_account_table(self):
        """
        更新账户表格
        """
        # 清空表格
        self.account_table.setRowCount(0)
        
        # 添加账户数据到表格
        for ccy, data in self.account_data.items():
            row_position = self.account_table.rowCount()
            self.account_table.insertRow(row_position)
            
            self.account_table.setItem(row_position, 0, QTableWidgetItem(ccy))
            self.account_table.setItem(row_position, 1, QTableWidgetItem(data['avail_balance']))
            self.account_table.setItem(row_position, 2, QTableWidgetItem(data['frozen_balance']))
            self.account_table.setItem(row_position, 3, QTableWidgetItem(data['total_balance']))

    def execute_trade(self):
        """
        执行交易
        """
        try:
            # 获取交易参数
            trade_pair = self.trade_pair_combo.currentText()
            trade_side = self.trade_side_combo.currentText()
            trade_type = self.trade_type_combo.currentText()
            price = self.trade_price_input.text()
            amount = self.trade_amount_input.text()

            # 验证参数
            if not price or not amount:
                self.statusBar.showMessage("请填写价格和数量")
                return

            # 模拟交易执行
            self.statusBar.showMessage(f"执行交易: {trade_side} {amount} {trade_pair} @ {price}")
            logger.info(f"执行交易: {trade_side} {amount} {trade_pair} @ {price}")

            # 模拟订单数据
            order_id = f"order_{datetime.now().timestamp()}"
            order_data = {
                'order_id': order_id,
                'inst_id': trade_pair,
                'side': trade_side,
                'type': trade_type,
                'price': price,
                'size': amount,
                'status': 'filled',
                'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }

            # 更新订单数据
            self.order_data[order_id] = order_data
            self.update_order_table()

            QMessageBox.information(self, "交易成功", f"交易已执行: {trade_side} {amount} {trade_pair}")
        except Exception as e:
            logger.error(f"执行交易错误: {e}")
            QMessageBox.error(self, "错误", f"执行交易失败: {str(e)}")

    def execute_batch_trade(self):
        """
        执行批量交易
        """
        try:
            # 获取批量交易指令
            batch_orders = self.batch_trade_text.toPlainText().strip().split('\n')
            
            if not batch_orders or batch_orders[0] == '':
                self.statusBar.showMessage("请输入批量交易指令")
                return
            
            # 执行每个交易指令
            success_count = 0
            error_count = 0
            
            for order_str in batch_orders:
                if not order_str.strip():
                    continue
                
                try:
                    # 解析交易指令
                    parts = order_str.strip().split(',')
                    if len(parts) != 5:
                        error_count += 1
                        continue
                    
                    trade_pair, trade_side, trade_type, price, amount = parts
                    
                    # 模拟交易执行
                    self.statusBar.showMessage(f"执行批量交易: {trade_side} {amount} {trade_pair} @ {price}")
                    logger.info(f"执行批量交易: {trade_side} {amount} {trade_pair} @ {price}")
                    
                    # 模拟订单数据
                    order_id = f"order_{datetime.now().timestamp()}_{success_count}"
                    order_data = {
                        'order_id': order_id,
                        'inst_id': trade_pair,
                        'side': trade_side,
                        'type': trade_type,
                        'price': price,
                        'size': amount,
                        'status': 'filled',
                        'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    }
                    
                    # 更新订单数据
                    self.order_data[order_id] = order_data
                    success_count += 1
                except Exception as e:
                    logger.error(f"执行批量交易错误: {e}")
                    error_count += 1
            
            # 更新订单表格
            self.update_order_table()
            
            # 显示执行结果
            QMessageBox.information(self, "批量交易完成", f"成功执行: {success_count} 笔\n失败: {error_count} 笔")
        except Exception as e:
            logger.error(f"执行批量交易错误: {e}")
            QMessageBox.error(self, "错误", f"执行批量交易失败: {str(e)}")

    def execute_conditional_trade(self):
        """
        执行条件单
        """
        try:
            # 获取条件单参数
            trade_pair = self.cond_trade_pair_combo.currentText()
            trade_side = self.cond_trade_side_combo.currentText()
            cond_type = self.cond_type_combo.currentText()
            trigger_price = self.cond_trigger_price_input.text()
            execute_price = self.cond_execute_price_input.text()
            amount = self.cond_amount_input.text()

            # 验证参数
            if not trigger_price or not execute_price or not amount:
                self.statusBar.showMessage("请填写触发价格、执行价格和数量")
                return

            # 模拟条件单创建
            self.statusBar.showMessage(f"创建条件单: {cond_type} {trade_side} {amount} {trade_pair} @ 触发价: {trigger_price}, 执行价: {execute_price}")
            logger.info(f"创建条件单: {cond_type} {trade_side} {amount} {trade_pair} @ 触发价: {trigger_price}, 执行价: {execute_price}")

            # 模拟条件单数据
            order_id = f"cond_order_{datetime.now().timestamp()}"
            order_data = {
                'order_id': order_id,
                'inst_id': trade_pair,
                'side': trade_side,
                'type': 'conditional',
                'price': execute_price,
                'size': amount,
                'status': 'pending',
                'update_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'cond_type': cond_type,
                'trigger_price': trigger_price,
            }

            # 更新订单数据
            self.order_data[order_id] = order_data
            self.update_order_table()

            QMessageBox.information(self, "条件单创建成功", f"条件单已创建: {cond_type} {trade_side} {amount} {trade_pair}")
        except Exception as e:
            logger.error(f"创建条件单错误: {e}")
            QMessageBox.error(self, "错误", f"创建条件单失败: {str(e)}")

    def init_monitor_tab(self):
        """
        初始化监控与分析标签页
        """
        monitor_layout = QVBoxLayout(self.monitor_tab)
        monitor_layout.setContentsMargins(15, 15, 15, 15)
        monitor_layout.setSpacing(15)

        # 系统状态监控
        system_status_group = QGroupBox("系统状态监控")
        system_status_layout = QVBoxLayout(system_status_group)
        system_status_layout.setContentsMargins(15, 15, 15, 15)

        # 系统状态信息
        status_info = QGridLayout()
        status_info.setSpacing(10)

        # 连接状态
        status_info.addWidget(QLabel("WebSocket连接:"), 0, 0)
        self.ws_status_label = QLabel("未连接")
        self.ws_status_label.setStyleSheet("font-weight: bold; color: red;")
        status_info.addWidget(self.ws_status_label, 0, 1)

        # 市场数据智能体状态
        status_info.addWidget(QLabel("市场数据智能体:"), 1, 0)
        self.agent_status_label = QLabel("未启动")
        self.agent_status_label.setStyleSheet("font-weight: bold; color: red;")
        status_info.addWidget(self.agent_status_label, 1, 1)

        # 策略运行状态
        status_info.addWidget(QLabel("运行中策略:"), 2, 0)
        self.strategy_count_label = QLabel("0个")
        self.strategy_count_label.setStyleSheet("font-weight: bold;")
        status_info.addWidget(self.strategy_count_label, 2, 1)

        # 系统健康状态
        status_info.addWidget(QLabel("系统健康状态:"), 3, 0)
        self.system_health_label = QLabel("正常")
        self.system_health_label.setStyleSheet("font-weight: bold; color: green;")
        status_info.addWidget(self.system_health_label, 3, 1)

        # 系统资源使用
        status_info.addWidget(QLabel("CPU使用率:"), 0, 2)
        self.cpu_usage_label = QLabel("0%")
        status_info.addWidget(self.cpu_usage_label, 0, 3)

        status_info.addWidget(QLabel("内存使用率:"), 1, 2)
        self.memory_usage_label = QLabel("0%")
        status_info.addWidget(self.memory_usage_label, 1, 3)

        system_status_layout.addLayout(status_info)
        
        # 健康检查按钮
        health_check_button = QPushButton("执行健康检查")
        health_check_button.clicked.connect(self.perform_health_check)
        system_status_layout.addWidget(health_check_button)
        
        monitor_layout.addWidget(system_status_group)
        
        # 启动系统健康检查定时器
        self.start_health_check_timer()

        # 交易统计
        trade_stats_group = QGroupBox("交易统计")
        trade_stats_layout = QHBoxLayout(trade_stats_group)
        trade_stats_layout.setContentsMargins(15, 15, 15, 15)
        trade_stats_layout.setSpacing(15)

        # 今日交易
        today_trades = QGroupBox("今日交易")
        today_trades_layout = QVBoxLayout(today_trades)
        self.today_trades_label = QLabel("0笔")
        self.today_trades_label.setStyleSheet("font-size: 24px; font-weight: bold;")
        today_trades_layout.addWidget(self.today_trades_label)
        today_trades_layout.addWidget(QLabel("笔数"))
        trade_stats_layout.addWidget(today_trades)

        # 胜率
        win_rate = QGroupBox("胜率")
        win_rate_layout = QVBoxLayout(win_rate)
        self.win_rate_label = QLabel("0%")
        self.win_rate_label.setStyleSheet("font-size: 24px; font-weight: bold;")
        win_rate_layout.addWidget(self.win_rate_label)
        win_rate_layout.addWidget(QLabel("百分比"))
        trade_stats_layout.addWidget(win_rate)

        # 总盈亏
        total_pnl = QGroupBox("总盈亏")
        total_pnl_layout = QVBoxLayout(total_pnl)
        self.total_pnl_label = QLabel("$0.00")
        self.total_pnl_label.setStyleSheet("font-size: 24px; font-weight: bold; color: green;")
        total_pnl_layout.addWidget(self.total_pnl_label)
        total_pnl_layout.addWidget(QLabel("美元"))
        trade_stats_layout.addWidget(total_pnl)

        monitor_layout.addWidget(trade_stats_group)

        # 日志输出
        log_group = QGroupBox("系统日志")
        log_layout = QVBoxLayout(log_group)
        log_layout.setContentsMargins(15, 15, 15, 15)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("font-family: Consolas; font-size: 12px;")
        log_layout.addWidget(self.log_text)
        monitor_layout.addWidget(log_group)

        # 设置拉伸比例
        monitor_layout.setStretch(0, 0)  # 系统状态不拉伸
        monitor_layout.setStretch(1, 0)  # 交易统计不拉伸
        monitor_layout.setStretch(2, 1)  # 日志输出拉伸

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = WebSocketGUI()
    window.show()
    sys.exit(app.exec_())